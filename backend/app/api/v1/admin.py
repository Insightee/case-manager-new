from __future__ import annotations

from typing import Optional

import csv
import io
import secrets
from datetime import date, datetime, timedelta, timezone

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, Request, Response, status
from pydantic import BaseModel, Field
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.api.deps import get_current_user, get_request_meta
from app.core.audit import log_audit
from app.core.config import settings
from app.core.database import get_db
from app.core.pagination import paginate_query, paginated_response
from app.schemas.pagination import PaginatedList
from app.core.module_access import (
    get_allowed_case_product_modules,
    get_user_features,
    modules_for_api,
    user_has_feature,
    validate_module_assignments,
)
from app.core.module_write import (
    ensure_billing_write_access,
    ensure_case_write_access,
    ensure_feature_write_access,
    ensure_product_module_write_access,
    guard_clinical_case,
)
from app.core.modules import legacy_role_defaults_flat, role_defaults_for_api
from app.core.permissions import case_scope_check, require_mutation_permission, require_permission, user_has_permission
from app.models.assignment import CaseAssignment, CaseAssignmentStatus
from app.models.case import Case, CaseStatus
from app.models.child import Child
from app.models.daily_log import DailyLog
from app.models.invoice import Invoice, InvoiceStatus
from app.models.report import MonthlyReport, ObservationReport, ReportStatus
from app.schemas.admin_reports import (
    AdminReportDetail,
    AdminReportListItem,
    AdminReportSummary,
    BulkReportAction,
    BulkReportResult,
    CmReviewAction,
    ReportCommentAction,
    ReportDocumentCommentCreate,
    SendForReviewAction,
)
from app.schemas.report import MonthlyReportUpdate, ObservationReportUpdate
from app.services import admin_report_service as admin_report_svc
from app.models.session import Session as TherapySession
from app.models.support_ticket import SupportTicket, TicketStatus
from app.models.therapist_profile import TherapistProfile, TherapistProfileStatus
from app.models.user import InviteToken, User
from app.schemas.admin_case_pipeline import AdminCasePipelineBoard
from app.schemas.admin_iep import AdminIepDashboard
from app.schemas.clinical import ObservationChecklistReview
from app.schemas.iep_plan import IepPlanSave, IepPlanSuggestionCreate
from app.schemas.therapist_onboarding import (
    TherapistBulkOnboardRequest,
    TherapistOnboardCreate,
    TherapistOnboardResult,
)
from app.models.service_category import ServiceCategory
from app.schemas.therapist_profile import (
    ServiceCategoryCreate,
    ServiceCategoryRead,
    ServiceCategoryUpdate,
    TherapistProfileAdminCreate,
    TherapistProfileRead,
    TherapistProfileReview,
    TherapistProfileUpdate,
)
from app.services import admin_case_pipeline_service as case_pipeline_svc
from app.services import admin_iep_service as admin_iep_svc
from app.services import therapist_onboarding_service as therapist_onboard_svc
from app.schemas.therapist_review import (
    TherapistReviewSummary,
    TherapistReviewsResponse,
    TherapistSessionReviewRead,
)
from app.schemas.allotment import CaseAllotRequest, ChildCreate, FamilyCreate
from app.schemas.rbac import RbacPreviewRequest
from app.schemas.user import InviteCreate, UserCreate, UserRead, UserUpdate
from app.core.rbac_access import (
    ASSIGNABLE_STAFF_ROLES,
    DEPRECATED_STAFF_ROLES,
    validate_assignable_staff_roles,
    module_catalog_entries,
    preview_access,
    sync_user_access_fields,
)
from app.services import auth_service, case_service, log_service, therapist_profile_service as profile_svc
from app.services.admin_scope_service import apply_case_scope
from app.services import therapist_review_service as review_svc
from app.core.permissions import RoleName

router = APIRouter(prefix="/admin", tags=["admin"])


def _admin_dashboard_user(user: User = Depends(get_current_user)) -> User:
    if not (
        user_has_permission(user, "case.read.all")
        or user_has_permission(user, "case.read.team")
        or user_has_permission(user, "admin.override")
    ):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient permissions")
    return user


def _case_filter(allowed: set[str] | None):
    if allowed is None:
        return []
    if not allowed:
        return [Case.id == -1]
    return [Case.product_module.in_(allowed)]


@router.get("/modules")
def list_product_modules(
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.core.rbac_access import module_catalog_entries

    from app.core.rbac_access import rbac_catalog_payload

    payload = rbac_catalog_payload(db)
    payload["role_defaults_flat"] = legacy_role_defaults_flat(db)
    return payload


def _user_to_read(u: User) -> UserRead:
    return UserRead(
        id=u.id,
        email=u.email,
        full_name=u.full_name,
        phone=u.phone,
        is_active=u.is_active,
        is_view_only=getattr(u, "is_view_only", False),
        roles=u.role_names,
        region=u.region,
        module_assignments=u.module_assignments or [],
        module_access_grants=getattr(u, "module_access_grants", None) or {},
        service_access_grants=getattr(u, "service_access_grants", None) or {},
        org_capability_grants=getattr(u, "org_capability_grants", None) or {},
        feature_overrides=getattr(u, "feature_overrides", None) or {},
    )


def _ensure_assignable_roles(role_names: list[str]) -> None:
    try:
        validate_assignable_staff_roles(role_names)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


def _apply_access_payload(
    user: User,
    *,
    role_names: list[str],
    module_assignments: list[str] | None = None,
    module_access_grants: dict | None = None,
    service_access_grants: dict | None = None,
    org_capability_grants: dict | None = None,
    feature_overrides: dict | None = None,
    view_only: bool | None = None,
    db: Session,
) -> None:
    if module_assignments is not None and module_access_grants is None and service_access_grants is None:
        validate_module_assignments(role_names, module_assignments, db)
    sync_user_access_fields(
        user,
        role_names=role_names,
        module_assignments=module_assignments,
        module_access_grants=module_access_grants,
        service_access_grants=service_access_grants,
        org_capability_grants=org_capability_grants,
        feature_overrides=feature_overrides,
        view_only=view_only,
        db=db,
    )
    roles_upper = {r.upper() for r in role_names}
    if "SUPER_ADMIN" not in roles_upper:
        validate_module_assignments(role_names, user.module_assignments, db)


@router.get("/rbac/catalog")
def rbac_catalog(
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.core.rbac_access import rbac_catalog_payload

    return {
        "assignable_roles": list(ASSIGNABLE_STAFF_ROLES),
        "deprecated_roles": sorted(DEPRECATED_STAFF_ROLES),
        **rbac_catalog_payload(db),
    }


@router.post("/rbac/preview")
def rbac_preview(
    payload: RbacPreviewRequest,
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    return preview_access(
        role_names=payload.role_names,
        module_access_grants=payload.module_access_grants,
        feature_overrides=payload.feature_overrides,
        view_only=payload.view_only,
        db=db,
    )


@router.get("/service-categories", response_model=list[ServiceCategoryRead])
def list_service_categories(
    include_inactive: bool = False,
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.services.service_category_service import category_to_read_dict

    stmt = select(ServiceCategory).order_by(ServiceCategory.sort_order, ServiceCategory.label)
    if not include_inactive:
        stmt = stmt.where(ServiceCategory.is_active.is_(True))
    rows = db.scalars(stmt).all()
    return [ServiceCategoryRead(**category_to_read_dict(r)) for r in rows]


@router.post("/service-categories", response_model=ServiceCategoryRead, status_code=status.HTTP_201_CREATED)
def create_service_category(
    payload: ServiceCategoryCreate,
    request: Request,
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    import re

    from app.services.service_category_service import (
        category_to_read_dict,
        normalize_product_modules_payload,
    )

    slug = payload.id
    if not slug:
        slug = re.sub(r"[^a-z0-9]+", "_", payload.label.strip().lower()).strip("_")
    try:
        product_modules = normalize_product_modules_payload(
            payload.product_modules,
            default_id=slug,
            default_label=payload.label.strip(),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    existing = db.get(ServiceCategory, slug)
    if existing:
        if not existing.is_active:
            existing.is_active = True
            existing.label = payload.label
            existing.description = payload.description or ""
            existing.sort_order = payload.sort_order
            existing.product_modules = product_modules
            db.commit()
            db.refresh(existing)
            return ServiceCategoryRead(**category_to_read_dict(existing))
        raise HTTPException(status_code=400, detail=f"Service category '{slug}' already exists")
    cat = ServiceCategory(
        id=slug,
        label=payload.label,
        description=payload.description or "",
        sort_order=payload.sort_order,
        product_modules=product_modules,
        is_active=True,
    )
    db.add(cat)
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="create", entity_type="service_category", entity_id=slug, **meta)
    db.commit()
    db.refresh(cat)
    return ServiceCategoryRead(**category_to_read_dict(cat))


@router.patch("/service-categories/{category_id}", response_model=ServiceCategoryRead)
def update_service_category(
    category_id: str,
    payload: ServiceCategoryUpdate,
    request: Request,
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.services.service_category_service import category_to_read_dict, normalize_product_modules_payload

    cat = db.get(ServiceCategory, category_id)
    if not cat:
        raise HTTPException(status_code=404, detail="Service category not found")
    if payload.label is not None:
        cat.label = payload.label
    if payload.description is not None:
        cat.description = payload.description
    if payload.sort_order is not None:
        cat.sort_order = payload.sort_order
    if payload.is_active is not None:
        cat.is_active = payload.is_active
    if payload.product_modules is not None:
        try:
            cat.product_modules = normalize_product_modules_payload(
                payload.product_modules,
                default_id=cat.id,
                default_label=cat.label,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="update", entity_type="service_category", entity_id=category_id, **meta)
    db.commit()
    db.refresh(cat)
    return ServiceCategoryRead(**category_to_read_dict(cat))


@router.delete("/service-categories/{category_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_service_category(
    category_id: str,
    request: Request,
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    cat = db.get(ServiceCategory, category_id)
    if not cat:
        raise HTTPException(status_code=404, detail="Service category not found")
    cat.is_active = False
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="delete", entity_type="service_category", entity_id=category_id, **meta)
    db.commit()


class ServiceProductCreate(BaseModel):
    name: str
    billing_model: str = "PER_SESSION"
    price_inr: Optional[float] = None
    package_sessions: Optional[int] = None
    discount_percent: Optional[float] = None
    total_inr: Optional[float] = None
    taxable: bool = True
    gst_rate_percent: Optional[float] = None
    gst_split: Optional[str] = None
    leave_policy: Optional[str] = None
    active: bool = True
    sort_order: int = 0


class ServiceProductUpdate(BaseModel):
    name: Optional[str] = None
    billing_model: Optional[str] = None
    price_inr: Optional[float] = None
    package_sessions: Optional[int] = None
    discount_percent: Optional[float] = None
    total_inr: Optional[float] = None
    taxable: Optional[bool] = None
    gst_rate_percent: Optional[float] = None
    gst_split: Optional[str] = None
    leave_policy: Optional[str] = None
    active: Optional[bool] = None
    sort_order: Optional[int] = None


@router.get("/service-categories/{category_id}/products")
def list_service_products(
    category_id: str,
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.services.service_product_service import list_products_for_category

    return list_products_for_category(db, category_id)


@router.post("/service-categories/{category_id}/products", status_code=status.HTTP_201_CREATED)
def create_service_product(
    category_id: str,
    payload: ServiceProductCreate,
    request: Request,
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.services.service_product_service import create_product

    try:
        row = create_product(db, category_id, payload.model_dump())
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    meta = get_request_meta(request)
    log_audit(
        db,
        actor_user_id=user.id,
        action="create",
        entity_type="service_product",
        entity_id=row["id"],
        **meta,
    )
    db.commit()
    return row


@router.patch("/service-products/{product_id}")
def update_service_product(
    product_id: int,
    payload: ServiceProductUpdate,
    request: Request,
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.services.service_product_service import update_product

    try:
        row = update_product(db, product_id, payload.model_dump(exclude_unset=True))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="update", entity_type="service_product", entity_id=product_id, **meta)
    db.commit()
    return row


@router.delete("/service-products/{product_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_service_product(
    product_id: int,
    request: Request,
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.services.service_product_service import delete_product

    try:
        delete_product(db, product_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="delete", entity_type="service_product", entity_id=product_id, **meta)
    db.commit()


class PrimaryCmUpdate(BaseModel):
    primary_case_manager_user_id: int
    mentor_user_id: Optional[int] = None
    update_active_cases: bool = False


@router.patch("/therapists/{user_id}/primary-cm")
def update_therapist_primary_cm(
    user_id: int,
    payload: PrimaryCmUpdate,
    request: Request,
    user: User = Depends(require_mutation_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.models.therapist_profile import TherapistProfile

    target = db.get(User, user_id)
    if not target or "THERAPIST" not in target.role_names:
        raise HTTPException(status_code=404, detail="Therapist not found")
    cm = db.get(User, payload.primary_case_manager_user_id)
    if not cm:
        raise HTTPException(status_code=400, detail="Case manager not found")
    if not {"CASE_MANAGER", "MODULE_ADMIN", "ADMIN", "SUPER_ADMIN"} & set(cm.role_names):
        raise HTTPException(status_code=400, detail="Selected user cannot be a primary case manager")
    profile = db.scalars(select(TherapistProfile).where(TherapistProfile.user_id == user_id)).first()
    if not profile:
        raise HTTPException(status_code=404, detail="Therapist profile not found")
    old_cm = profile.supervisor_user_id
    profile.supervisor_user_id = payload.primary_case_manager_user_id
    if payload.mentor_user_id is not None:
        profile.mentor_user_id = payload.mentor_user_id
    if payload.update_active_cases and old_cm:
        from app.models.case import Case, CaseStatus

        cases = db.scalars(
            select(Case).where(
                Case.status == CaseStatus.ACTIVE,
                Case.case_manager_user_id == old_cm,
            )
        ).all()
        for case in cases:
            case.case_manager_user_id = payload.primary_case_manager_user_id
    meta = get_request_meta(request)
    log_audit(
        db,
        actor_user_id=user.id,
        action="change_primary_cm",
        entity_type="therapist_profile",
        entity_id=profile.id,
        old_value={"primary_cm_user_id": old_cm},
        new_value={
            "primary_cm_user_id": payload.primary_case_manager_user_id,
            "update_active_cases": payload.update_active_cases,
        },
        **meta,
    )
    db.commit()
    return {"user_id": user_id, "primary_case_manager_user_id": profile.supervisor_user_id}


def _workbench_user(user: User = Depends(get_current_user)) -> User:
    if not (
        user_has_permission(user, "case.read.team")
        and (
            user_has_permission(user, "monthly_report.approve")
            or user_has_permission(user, "daily_log.review")
        )
    ):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient permissions")
    return user


@router.get("/workbench/summary")
def workbench_summary(
    user: User = Depends(_workbench_user),
    db: Session = Depends(get_db),
):
    from app.services import admin_workbench_service as workbench_svc

    return workbench_svc.build_workbench_summary(db, user)


@router.get("/home")
def admin_home(
    user: User = Depends(_admin_dashboard_user),
    db: Session = Depends(get_db),
):
    from app.core.config import settings
    from app.core.database import ensure_sqlite_schema_patches
    from app.services import admin_home_service

    if settings.is_sqlite:
        ensure_sqlite_schema_patches()

    return admin_home_service.build_admin_home(db, user)


def _require_case_manager_home(user: User = Depends(get_current_user)) -> User:
    if "CASE_MANAGER" not in user.role_names:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Case Manager home requires CASE_MANAGER role",
        )
    if not user_has_permission(user, "case.read.team"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient permissions")
    return user


@router.get("/cm/home")
def admin_cm_home(
    user: User = Depends(_require_case_manager_home),
    db: Session = Depends(get_db),
):
    from app.schemas.admin_cm_home import AdminCmHomeResponse
    from app.services import admin_cm_home_service

    return AdminCmHomeResponse.model_validate(admin_cm_home_service.build_cm_home(db, user))


@router.get("/audit")
def admin_audit_list(
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    case_id: Optional[int] = None,
    limit: int = Query(50, ge=1, le=100),
    cursor: Optional[int] = None,
    user: User = Depends(_admin_dashboard_user),
    db: Session = Depends(get_db),
):
    from app.services import audit_service

    try:
        return audit_service.list_audit_events(
            db,
            user,
            entity_type=entity_type,
            entity_id=entity_id,
            case_id=case_id,
            limit=limit,
            cursor=cursor,
        )
    except PermissionError as e:
        raise HTTPException(status_code=403, detail=str(e)) from e


@router.get("/audit/app-usage-summary")
def admin_app_usage_summary(
    days: int = Query(7, ge=1, le=90),
    portal: Optional[str] = Query(None),
    user_id: Optional[int] = Query(None),
    user: User = Depends(_admin_dashboard_user),
    db: Session = Depends(get_db),
):
    if not user_has_permission(user, "admin.override"):
        raise HTTPException(status_code=403, detail="Super admin permission required")
    from app.services import audit_service

    end_at = datetime.now(timezone.utc)
    start_at = end_at - timedelta(days=days)
    try:
        return audit_service.app_usage_summary(
            db,
            user,
            start_at=start_at,
            end_at=end_at,
            portal=portal,
            staff_user_id=user_id,
        )
    except PermissionError as e:
        raise HTTPException(status_code=403, detail=str(e)) from e


@router.get("/cases/{case_id}/timeline")
def admin_case_timeline(
    case_id: int,
    limit: int = Query(40, ge=1, le=100),
    user: User = Depends(_admin_dashboard_user),
    db: Session = Depends(get_db),
):
    from app.services import audit_service

    try:
        return {"items": audit_service.case_timeline(db, user, case_id, limit=limit)}
    except PermissionError as e:
        raise HTTPException(status_code=403, detail=str(e)) from e


@router.get("/dashboard/summary")
def dashboard_summary(
    user: User = Depends(_admin_dashboard_user),
    db: Session = Depends(get_db),
):
    from app.core.config import settings
    from app.core.database import ensure_sqlite_schema_patches

    if settings.is_sqlite:
        ensure_sqlite_schema_patches()

    allowed_cases = get_allowed_case_product_modules(user)

    def _count_case_status(status: CaseStatus) -> int:
        stmt = select(func.count()).select_from(Case).where(Case.status == status)
        stmt = apply_case_scope(stmt, user)
        return db.scalar(stmt) or 0

    pending_stmt = (
        select(Case.id, Case.case_code, Case.service_type, Case.status, Child.first_name, Child.last_name)
        .join(Child, Case.child_id == Child.id)
        .where(Case.status == CaseStatus.PENDING_ALLOTMENT)
    )
    pending_stmt = apply_case_scope(pending_stmt, user).order_by(Case.created_at.desc()).limit(6)
    pending_rows = db.execute(pending_stmt).all()

    report_stmt = (
        select(
            MonthlyReport.id,
            MonthlyReport.case_id,
            MonthlyReport.month,
            MonthlyReport.status,
            Case.case_code,
            Child.first_name,
            Child.last_name,
        )
        .join(Case, MonthlyReport.case_id == Case.id)
        .join(Child, Case.child_id == Child.id)
        .where(MonthlyReport.status == ReportStatus.UNDER_REVIEW)
    )
    report_stmt = apply_case_scope(report_stmt, user).order_by(MonthlyReport.updated_at.desc()).limit(6)
    report_rows = db.execute(report_stmt).all() if user_has_feature(user, "reports") else []

    invoice_rows = []
    invoices_pending = 0
    if user_has_feature(user, "invoices"):
        scoped_therapist_ids = None
        if allowed_cases is not None:
            if not allowed_cases:
                scoped_therapist_ids = []
            else:
                scoped_therapist_ids = list(
                    db.scalars(
                        select(CaseAssignment.therapist_user_id)
                        .join(Case, CaseAssignment.case_id == Case.id)
                        .where(
                            CaseAssignment.status == CaseAssignmentStatus.ACTIVE,
                            Case.product_module.in_(allowed_cases),
                        )
                        .distinct()
                    ).all()
                )
        invoice_base = select(Invoice).where(Invoice.status == InvoiceStatus.IN_REVIEW)
        if scoped_therapist_ids is not None:
            if not scoped_therapist_ids:
                invoice_base = invoice_base.where(Invoice.id == -1)
            else:
                invoice_base = invoice_base.where(Invoice.therapist_user_id.in_(scoped_therapist_ids))
        invoice_rows = list(
            db.scalars(invoice_base.order_by(Invoice.updated_at.desc()).limit(6)).all()
        )
        count_stmt = select(func.count()).select_from(Invoice).where(Invoice.status == InvoiceStatus.IN_REVIEW)
        if scoped_therapist_ids is not None:
            if not scoped_therapist_ids:
                count_stmt = count_stmt.where(Invoice.id == -1)
            else:
                count_stmt = count_stmt.where(Invoice.therapist_user_id.in_(scoped_therapist_ids))
        invoices_pending = db.scalar(count_stmt) or 0

    ticket_filters = []
    if allowed_cases is not None:
        if allowed_cases:
            ticket_filters.append(
                or_(SupportTicket.product_module.in_(allowed_cases), SupportTicket.product_module.is_(None))
            )
        else:
            ticket_filters.append(SupportTicket.id == -1)

    ticket_rows = []
    open_tickets = 0
    if user_has_feature(user, "tickets"):
        ticket_rows = list(
            db.scalars(
                select(SupportTicket)
                .where(SupportTicket.status == TicketStatus.OPEN, *ticket_filters)
                .order_by(SupportTicket.updated_at.desc())
                .limit(6)
            ).all()
        )
        open_tickets = (
            db.scalar(
                select(func.count()).select_from(SupportTicket).where(SupportTicket.status == TicketStatus.OPEN, *ticket_filters)
            )
            or 0
        )

    reports_in_review = 0
    if user_has_feature(user, "reports"):
        reports_count_stmt = (
            select(func.count())
            .select_from(MonthlyReport)
            .join(Case, MonthlyReport.case_id == Case.id)
            .where(MonthlyReport.status == ReportStatus.UNDER_REVIEW)
        )
        reports_in_review = db.scalar(apply_case_scope(reports_count_stmt, user)) or 0

    total_cases_stmt = apply_case_scope(select(func.count()).select_from(Case), user)

    from app.services import admin_workbench_service as wb_svc

    ops_counts = wb_svc.build_ops_counts(db, user)

    return {
        "open_cases": _count_case_status(CaseStatus.ACTIVE),
        "pending_allotment": _count_case_status(CaseStatus.PENDING_ALLOTMENT),
        "suspended_cases": _count_case_status(CaseStatus.SUSPENDED),
        "closed_cases": _count_case_status(CaseStatus.CLOSED),
        "total_cases": db.scalar(total_cases_stmt) or 0,
        "reports_in_review": reports_in_review,
        "invoices_pending": invoices_pending,
        "open_tickets": open_tickets,
        "observation_checklists_pending": ops_counts.get("observation_checklists_pending", 0),
        "observation_checklists_overdue": ops_counts.get("observation_checklists_overdue", 0),
        "observation_reports_in_review": ops_counts.get("observation_reports_in_review", 0),
        "status_requests_pending": ops_counts.get("status_requests_pending", 0),
        "client_payments_pending_review": ops_counts.get("client_payments_pending_review", 0),
        "iep_attention": ops_counts.get("iep_attention", 0),
        "iep_plans_draft": ops_counts.get("iep_plans_draft", 0),
        "status_breakdown": {
            "ACTIVE": _count_case_status(CaseStatus.ACTIVE),
            "PENDING_ALLOTMENT": _count_case_status(CaseStatus.PENDING_ALLOTMENT),
            "SUSPENDED": _count_case_status(CaseStatus.SUSPENDED),
            "CLOSED": _count_case_status(CaseStatus.CLOSED),
        },
        "pending_allotment_queue": [
            {
                "id": row.id,
                "case_code": row.case_code,
                "child_name": f"{row.first_name} {row.last_name}".strip(),
                "service_type": row.service_type,
                "status": row.status.value if hasattr(row.status, "value") else str(row.status),
            }
            for row in pending_rows
        ],
        "reports_queue": [
            {
                "id": row.id,
                "case_id": row.case_id,
                "month": row.month,
                "status": row.status.value if hasattr(row.status, "value") else str(row.status),
                "case_code": row.case_code,
                "child_name": f"{row.first_name} {row.last_name}".strip(),
            }
            for row in report_rows
        ],
        "invoices_queue": [
            {
                "id": inv.id,
                "month": inv.month,
                "amount_inr": float(inv.amount_inr),
                "status": inv.status.value,
                "therapist_user_id": inv.therapist_user_id,
            }
            for inv in invoice_rows
        ],
        "tickets_queue": [
            {
                "id": t.id,
                "subject": t.subject,
                "status": t.status.value,
                "product_module": t.product_module,
            }
            for t in ticket_rows
        ],
    }


@router.get("/sessions/analytics")
def sessions_analytics(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    therapist_id: Optional[int] = None,
    product_module: Optional[str] = None,
    case_id: Optional[int] = None,
    session_status: Optional[str] = Query(None, alias="status"),
    user: User = Depends(_admin_dashboard_user),
    db: Session = Depends(get_db),
):
    from datetime import date, timedelta
    from sqlalchemy import case as sa_case, cast, String, distinct
    from app.models.session import SessionStatus

    today = date.today()
    d_from = date.fromisoformat(date_from) if date_from else today - timedelta(days=29)
    d_to = date.fromisoformat(date_to) if date_to else today

    allowed_cases = get_allowed_case_product_modules(user)
    base_filters = []
    if allowed_cases is not None:
        base_filters.append(Case.product_module.in_(allowed_cases) if allowed_cases else Case.id == -1)
    if therapist_id:
        base_filters.append(TherapySession.therapist_user_id == therapist_id)
    if product_module:
        base_filters.append(Case.product_module == product_module)
    if session_status:
        base_filters.append(TherapySession.status == session_status)
    if case_id:
        base_filters.append(TherapySession.case_id == case_id)

    # Base join: session → case
    base_q = (
        select(TherapySession)
        .join(Case, TherapySession.case_id == Case.id)
        .where(
            TherapySession.scheduled_date >= d_from,
            TherapySession.scheduled_date <= d_to,
            *base_filters,
        )
    )

    # Today count
    today_count = db.scalar(
        select(func.count())
        .select_from(TherapySession)
        .join(Case, TherapySession.case_id == Case.id)
        .where(TherapySession.scheduled_date == today, *base_filters)
    ) or 0

    # This-week count
    week_start = today - timedelta(days=today.weekday())
    week_count = db.scalar(
        select(func.count())
        .select_from(TherapySession)
        .join(Case, TherapySession.case_id == Case.id)
        .where(TherapySession.scheduled_date >= week_start, TherapySession.scheduled_date <= today, *base_filters)
    ) or 0

    # Status counts
    status_rows = db.execute(
        select(TherapySession.status, func.count().label("n"))
        .join(Case, TherapySession.case_id == Case.id)
        .where(
            TherapySession.scheduled_date >= d_from,
            TherapySession.scheduled_date <= d_to,
            *base_filters,
        )
        .group_by(TherapySession.status)
    ).all()
    status_counts = {row.status.value if hasattr(row.status, "value") else str(row.status): row.n for row in status_rows}

    # By therapist
    therapist_rows = db.execute(
        select(
            TherapySession.therapist_user_id,
            User.full_name,
            func.count().label("total"),
            func.sum(sa_case((TherapySession.status == SessionStatus.COMPLETED, 1), else_=0)).label("completed"),
            func.sum(sa_case((TherapySession.status == SessionStatus.CANCELLED, 1), else_=0)).label("cancelled"),
            func.sum(sa_case((TherapySession.status == SessionStatus.SCHEDULED, 1), else_=0)).label("scheduled"),
        )
        .join(Case, TherapySession.case_id == Case.id)
        .outerjoin(User, TherapySession.therapist_user_id == User.id)
        .where(
            TherapySession.scheduled_date >= d_from,
            TherapySession.scheduled_date <= d_to,
            *base_filters,
        )
        .group_by(TherapySession.therapist_user_id, User.full_name)
        .order_by(func.count().desc())
    ).all()
    by_therapist = [
        {
            "therapist_id": r.therapist_user_id,
            "name": r.full_name or f"Therapist #{r.therapist_user_id}",
            "total": r.total,
            "completed": int(r.completed or 0),
            "cancelled": int(r.cancelled or 0),
            "scheduled": int(r.scheduled or 0),
        }
        for r in therapist_rows
    ]

    # By product module
    product_rows = db.execute(
        select(
            Case.product_module,
            func.count().label("total"),
            func.sum(sa_case((TherapySession.status == SessionStatus.COMPLETED, 1), else_=0)).label("completed"),
            func.sum(sa_case((TherapySession.status == SessionStatus.CANCELLED, 1), else_=0)).label("cancelled"),
        )
        .join(Case, TherapySession.case_id == Case.id)
        .where(
            TherapySession.scheduled_date >= d_from,
            TherapySession.scheduled_date <= d_to,
            *base_filters,
        )
        .group_by(Case.product_module)
        .order_by(func.count().desc())
    ).all()
    by_product = [
        {
            "module": r.product_module or "unknown",
            "total": r.total,
            "completed": int(r.completed or 0),
            "cancelled": int(r.cancelled or 0),
        }
        for r in product_rows
    ]

    # By day (within requested range)
    day_rows = db.execute(
        select(
            TherapySession.scheduled_date,
            func.count().label("total"),
            func.sum(sa_case((TherapySession.status == SessionStatus.COMPLETED, 1), else_=0)).label("completed"),
            func.sum(sa_case((TherapySession.status == SessionStatus.CANCELLED, 1), else_=0)).label("cancelled"),
        )
        .join(Case, TherapySession.case_id == Case.id)
        .where(
            TherapySession.scheduled_date >= d_from,
            TherapySession.scheduled_date <= d_to,
            *base_filters,
        )
        .group_by(TherapySession.scheduled_date)
        .order_by(TherapySession.scheduled_date)
    ).all()
    by_day = [
        {
            "date": r.scheduled_date.isoformat(),
            "total": r.total,
            "completed": int(r.completed or 0),
            "cancelled": int(r.cancelled or 0),
        }
        for r in day_rows
    ]

    # By month (last 12 months from today)
    month_start = (today.replace(day=1) - timedelta(days=365)).replace(day=1)
    from sqlalchemy import extract
    month_rows = db.execute(
        select(
            extract("year", TherapySession.scheduled_date).label("yr"),
            extract("month", TherapySession.scheduled_date).label("mo"),
            func.count().label("total"),
            func.sum(sa_case((TherapySession.status == SessionStatus.COMPLETED, 1), else_=0)).label("completed"),
            func.sum(sa_case((TherapySession.status == SessionStatus.CANCELLED, 1), else_=0)).label("cancelled"),
        )
        .join(Case, TherapySession.case_id == Case.id)
        .where(
            TherapySession.scheduled_date >= month_start,
            *([TherapySession.therapist_user_id == therapist_id] if therapist_id else []),
            *([Case.product_module == product_module] if product_module else []),
            *(base_filters[0:1] if base_filters else []),
        )
        .group_by("yr", "mo")
        .order_by("yr", "mo")
    ).all()
    by_month = [
        {
            "month": f"{int(r.yr)}-{int(r.mo):02d}",
            "total": r.total,
            "completed": int(r.completed or 0),
            "cancelled": int(r.cancelled or 0),
        }
        for r in month_rows
    ]

    # Recent sessions for the table (last 50 within filter range)
    from sqlalchemy.orm import selectinload
    recent_limit = 200 if case_id else 50
    sessions_q = (
        select(TherapySession)
        .join(Case, TherapySession.case_id == Case.id)
        .options(
            selectinload(TherapySession.case).selectinload(Case.child),
            selectinload(TherapySession.daily_log),
        )
        .where(
            TherapySession.scheduled_date >= d_from,
            TherapySession.scheduled_date <= d_to,
            *base_filters,
        )
        .order_by(TherapySession.scheduled_date.desc())
        .limit(recent_limit)
    )
    sessions_rows = db.scalars(sessions_q).all()
    recent_sessions = []
    for s in sessions_rows:
        case_obj = s.case
        child_name = (case_obj.child.full_name if case_obj and case_obj.child else None)
        therapist_name = None
        if s.therapist_user_id:
            tu = db.get(User, s.therapist_user_id)
            therapist_name = tu.full_name if tu else None
        duration_mins = None
        if s.actual_start_at and s.actual_end_at:
            duration_mins = int((s.actual_end_at - s.actual_start_at).total_seconds() / 60)
        elif s.start_time and s.end_time:
            from datetime import datetime as dt
            s_start = dt.combine(s.scheduled_date, s.start_time)
            s_end = dt.combine(s.scheduled_date, s.end_time)
            duration_mins = int((s_end - s_start).total_seconds() / 60)
        recent_sessions.append({
            "id": s.id,
            "case_id": s.case_id,
            "case_code": case_obj.case_code if case_obj else None,
            "child_name": child_name,
            "therapist_id": s.therapist_user_id,
            "therapist_name": therapist_name,
            "product_module": case_obj.product_module if case_obj else None,
            "scheduled_date": s.scheduled_date.isoformat(),
            "start_time": s.start_time.isoformat() if s.start_time else None,
            "end_time": s.end_time.isoformat() if s.end_time else None,
            "actual_start_at": s.actual_start_at.isoformat() if s.actual_start_at else None,
            "actual_end_at": s.actual_end_at.isoformat() if s.actual_end_at else None,
            "mode": s.mode.value if hasattr(s.mode, "value") else s.mode,
            "status": s.status.value if hasattr(s.status, "value") else s.status,
            "duration_mins": duration_mins,
            "has_daily_log": s.daily_log is not None,
        })

    return {
        "today_count": today_count,
        "week_count": week_count,
        "status_counts": status_counts,
        "by_therapist": by_therapist,
        "by_product": by_product,
        "by_day": by_day,
        "by_month": by_month,
        "recent_sessions": recent_sessions,
        "date_from": d_from.isoformat(),
        "date_to": d_to.isoformat(),
    }


@router.post("/sessions/{session_id}/flag")
def flag_session_for_review(
    session_id: int,
    payload: dict,
    request: Request,
    user: User = Depends(require_permission("case.read.all")),
    db: Session = Depends(get_db),
):
    session = db.get(TherapySession, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    case = case_service.get_case(db, session.case_id)
    if not case or not case_scope_check(db, user, case):
        raise HTTPException(status_code=403, detail="Access denied")

    reason = payload.get("reason", "Session flagged for review")
    notes = payload.get("notes", "")
    subject = f"Session #{session_id} flagged: {reason}"
    body = f"Session {session_id} on {session.scheduled_date} was flagged for review.\n\nReason: {reason}"
    if notes:
        body += f"\n\nNotes: {notes}"

    from app.models.support_ticket import TicketCategory, TicketTopic
    ticket = SupportTicket(
        case_id=session.case_id,
        raised_by_user_id=user.id,
        assigned_to_user_id=session.therapist_user_id,
        product_module=case.product_module if case else None,
        category=TicketCategory.SERVICE,
        topic=TicketTopic.THERAPIST,
        subject=subject,
        body=body,
        status=TicketStatus.OPEN,
    )
    db.add(ticket)
    db.flush()
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="flag_session", entity_type="session", entity_id=session_id, **meta)
    db.commit()
    return {"ticket_id": ticket.id, "subject": ticket.subject}


@router.get("/sessions/export/xlsx")
def export_sessions_xlsx(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    therapist_id: Optional[int] = None,
    product_module: Optional[str] = None,
    case_id: Optional[int] = None,
    session_status: Optional[str] = Query(None, alias="status"),
    user: User = Depends(_admin_dashboard_user),
    db: Session = Depends(get_db),
):
    import openpyxl
    from datetime import date, timedelta
    from io import BytesIO
    from sqlalchemy import case as sa_case
    from app.models.session import SessionStatus

    today = date.today()
    d_from = date.fromisoformat(date_from) if date_from else today - timedelta(days=29)
    d_to = date.fromisoformat(date_to) if date_to else today
    allowed_cases = get_allowed_case_product_modules(user)
    base_filters = []
    if allowed_cases is not None:
        base_filters.append(Case.product_module.in_(allowed_cases) if allowed_cases else Case.id == -1)
    if therapist_id:
        base_filters.append(TherapySession.therapist_user_id == therapist_id)
    if product_module:
        base_filters.append(Case.product_module == product_module)
    if session_status:
        base_filters.append(TherapySession.status == session_status)
    if case_id:
        base_filters.append(TherapySession.case_id == case_id)

    from sqlalchemy.orm import selectinload
    sessions_rows = db.scalars(
        select(TherapySession)
        .join(Case, TherapySession.case_id == Case.id)
        .options(selectinload(TherapySession.case).selectinload(Case.child))
        .where(TherapySession.scheduled_date >= d_from, TherapySession.scheduled_date <= d_to, *base_filters)
        .order_by(TherapySession.scheduled_date.desc())
    ).all()

    from app.services.export_document_service import export_meta, xlsx_footer_rows, xlsx_preamble_rows

    meta = export_meta(user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sessions"
    for row in xlsx_preamble_rows(
        "Session logs export",
        f"Period: {d_from.isoformat()} to {d_to.isoformat()}",
        meta,
    ):
        ws.append(row)
    headers = ["Session ID", "Date", "Start", "End", "Actual Start", "Actual End", "Duration (min)",
               "Case Code", "Child", "Therapist ID", "Product Module", "Mode", "Status"]
    ws.append(headers)

    for s in sessions_rows:
        case_obj = s.case
        child_name = (case_obj.child.full_name if case_obj and case_obj.child else "")
        duration_mins = ""
        if s.actual_start_at and s.actual_end_at:
            duration_mins = int((s.actual_end_at - s.actual_start_at).total_seconds() / 60)
        elif s.start_time and s.end_time:
            from datetime import datetime as dt
            s_start = dt.combine(s.scheduled_date, s.start_time)
            s_end = dt.combine(s.scheduled_date, s.end_time)
            duration_mins = int((s_end - s_start).total_seconds() / 60)
        ws.append([
            s.id,
            s.scheduled_date.isoformat(),
            s.start_time.isoformat() if s.start_time else "",
            s.end_time.isoformat() if s.end_time else "",
            s.actual_start_at.isoformat() if s.actual_start_at else "",
            s.actual_end_at.isoformat() if s.actual_end_at else "",
            duration_mins,
            case_obj.case_code if case_obj else "",
            child_name,
            s.therapist_user_id,
            case_obj.product_module if case_obj else "",
            s.mode.value if hasattr(s.mode, "value") else str(s.mode),
            s.status.value if hasattr(s.status, "value") else str(s.status),
        ])
    for row in xlsx_footer_rows(meta):
        ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=sessions_{d_from}_{d_to}.xlsx"},
    )


@router.get("/sessions/export/pdf")
def export_sessions_pdf(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    therapist_id: Optional[int] = None,
    product_module: Optional[str] = None,
    case_id: Optional[int] = None,
    session_status: Optional[str] = Query(None, alias="status"),
    user: User = Depends(_admin_dashboard_user),
    db: Session = Depends(get_db),
):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from datetime import date, timedelta
    from io import BytesIO
    from sqlalchemy.orm import selectinload

    today = date.today()
    d_from = date.fromisoformat(date_from) if date_from else today - timedelta(days=29)
    d_to = date.fromisoformat(date_to) if date_to else today
    allowed_cases = get_allowed_case_product_modules(user)
    base_filters = []
    if allowed_cases is not None:
        base_filters.append(Case.product_module.in_(allowed_cases) if allowed_cases else Case.id == -1)
    if therapist_id:
        base_filters.append(TherapySession.therapist_user_id == therapist_id)
    if product_module:
        base_filters.append(Case.product_module == product_module)
    if session_status:
        base_filters.append(TherapySession.status == session_status)
    if case_id:
        base_filters.append(TherapySession.case_id == case_id)

    sessions_rows = db.scalars(
        select(TherapySession)
        .join(Case, TherapySession.case_id == Case.id)
        .options(selectinload(TherapySession.case).selectinload(Case.child))
        .where(TherapySession.scheduled_date >= d_from, TherapySession.scheduled_date <= d_to, *base_filters)
        .order_by(TherapySession.scheduled_date.desc())
        .limit(500)
    ).all()

    from app.services.export_document_service import export_meta

    meta = export_meta(user)
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=40, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph("Session logs export", styles["Title"]))
    elements.append(Paragraph(f"Period: {d_from.isoformat()} to {d_to.isoformat()}", styles["Normal"]))
    elements.append(
        Paragraph(f"Generated by {meta['generated_by']} · {meta['generated_at']}", styles["Normal"])
    )
    elements.append(Spacer(1, 12))

    table_data = [["Date", "Case", "Child", "Therapist", "Module", "Mode", "Status", "Duration"]]
    for s in sessions_rows:
        c = s.case
        child_name = (c.child.full_name if c and c.child else "")
        duration = ""
        if s.actual_start_at and s.actual_end_at:
            duration = f"{int((s.actual_end_at - s.actual_start_at).total_seconds() / 60)} min"
        elif s.start_time and s.end_time:
            from datetime import datetime as dt
            duration = f"{int((dt.combine(s.scheduled_date, s.end_time) - dt.combine(s.scheduled_date, s.start_time)).total_seconds() / 60)} min"
        table_data.append([
            s.scheduled_date.isoformat(),
            c.case_code if c else "",
            child_name,
            str(s.therapist_user_id),
            c.product_module if c else "",
            s.mode.value if hasattr(s.mode, "value") else str(s.mode),
            s.status.value if hasattr(s.status, "value") else str(s.status),
            duration,
        ])

    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6366f1")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 16))
    elements.append(
        Paragraph(
            f"Document generated by {meta['generated_by']} on {meta['generated_at']}",
            styles["Normal"],
        )
    )
    doc.build(elements)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=sessions_{d_from}_{d_to}.pdf"},
    )


@router.get("/session-logs/summary")
def admin_session_logs_summary(
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    product_module: Optional[str] = None,
    user: User = Depends(_admin_dashboard_user),
    db: Session = Depends(get_db),
):
    from app.services import session_log_service

    if not (
        user_has_permission(user, "daily_log.review")
        or user_has_permission(user, "case.read.all")
        or user_has_permission(user, "case.read.team")
        or user_has_permission(user, "case.read.scoped")
    ):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    return session_log_service.admin_session_logs_summary(
        db,
        user,
        from_date=from_date,
        to_date=to_date,
        product_module=product_module,
    )


@router.get("/session-logs")
def admin_list_session_logs(
    status: Optional[str] = Query(None, description="missing|pending|submitted|approved"),
    case_id: Optional[int] = None,
    therapist_user_id: Optional[int] = None,
    product_module: Optional[str] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    user: User = Depends(_admin_dashboard_user),
    db: Session = Depends(get_db),
):
    from app.schemas.session_log_portal import AdminSessionLogListResponse, SessionLogRead
    from app.services import session_log_service

    if status and status not in ("missing", "pending", "submitted", "approved"):
        raise HTTPException(status_code=400, detail="Invalid status filter")
    if not (
        user_has_permission(user, "daily_log.review")
        or user_has_permission(user, "case.read.all")
        or user_has_permission(user, "case.read.team")
        or user_has_permission(user, "case.read.scoped")
    ):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    data = session_log_service.list_admin_session_logs(
        db,
        user,
        status=status,  # type: ignore[arg-type]
        case_id=case_id,
        therapist_user_id=therapist_user_id,
        product_module=product_module,
        from_date=from_date,
        to_date=to_date,
        page=page,
        page_size=page_size,
    )
    return AdminSessionLogListResponse(
        items=[SessionLogRead(**row) for row in data["items"]],
        total=data["total"],
        page=data["page"],
        page_size=data["page_size"],
        pages=data["pages"],
    )


@router.get("/session-logs/export")
def export_session_logs(
    therapist_user_id: Optional[int] = None,
    month: Optional[str] = None,
    product_module: Optional[str] = None,
    user: User = Depends(_admin_dashboard_user),
    db: Session = Depends(get_db),
):
    logs = log_service.list_logs(db, therapist_user_id=therapist_user_id, month=month, product_module=product_module)
    scoped = []
    for log in logs:
        if not log.session:
            continue
        case = case_service.get_case(db, log.session.case_id)
        if case and case_scope_check(db, user, case):
            scoped.append(log)
    logs = scoped
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "log_id", "session_id", "case_id", "attendance", "session_notes", "activities",
        "goals", "observations", "follow_ups", "parent_notes", "scheduled_date",
        "start_time", "end_time", "actual_start", "actual_end", "submitted_at", "approval_status", "late_addition",
    ])
    for log in logs:
        s = log.session
        writer.writerow([
            log.id,
            log.session_id,
            s.case_id if s else "",
            log.attendance_status,
            log.session_notes or "",
            log.activities_done or "",
            log.goals_addressed or "",
            log.observations or "",
            log.follow_ups or "",
            log.parent_notes or "",
            s.scheduled_date.isoformat() if s else "",
            s.start_time.isoformat() if s and s.start_time else "",
            s.end_time.isoformat() if s and s.end_time else "",
            s.actual_start_at.isoformat() if s and s.actual_start_at else "",
            s.actual_end_at.isoformat() if s and s.actual_end_at else "",
            log.submitted_at.isoformat() if log.submitted_at else "",
            log.approval_status.value,
            log.late_addition,
        ])
    return Response(content=output.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=session_logs.csv"})


def require_user_directory_read(user: User = Depends(get_current_user)) -> User:
    """Staff pickers (tickets, CM meetings) use therapist.read; People uses user.manage."""
    if user_has_permission(user, "user.manage") or user_has_permission(user, "therapist.read"):
        return user
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to list users")


@router.get("/users", response_model=PaginatedList[UserRead])
def list_users(
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=100),
    user: User = Depends(require_user_directory_read),
    db: Session = Depends(get_db),
):
    stmt = select(User).order_by(User.email)
    users, total = paginate_query(db, stmt, page=page, page_size=page_size)
    items = [_user_to_read(u) for u in users]
    return PaginatedList[UserRead](
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=max(1, (total + page_size - 1) // page_size),
    )


@router.get("/users/directory", response_model=list)
def users_directory(
    roles: Optional[str] = Query(None, description="Comma-separated role names e.g. THERAPIST,CASE_MANAGER"),
    search: Optional[str] = None,
    active_only: bool = True,
    limit: int = Query(500, ge=1, le=500),
    user: User = Depends(require_user_directory_read),
    db: Session = Depends(get_db),
):
    from app.schemas.user import UserDirectoryItem

    role_set = {r.strip().upper() for r in (roles or "").split(",") if r.strip()}
    stmt = select(User).order_by(User.full_name, User.email)
    if active_only:
        stmt = stmt.where(User.is_active.is_(True))
    rows = list(db.scalars(stmt.limit(limit)).all())
    out: list[UserDirectoryItem] = []
    q = (search or "").strip().lower()
    for u in rows:
        user_roles = [str(r).upper() for r in (u.role_names or [])]
        if role_set and not role_set.intersection(user_roles):
            continue
        if q:
            hay = f"{u.full_name} {u.email}".lower()
            if q not in hay:
                continue
        out.append(
            UserDirectoryItem(
                id=u.id,
                email=u.email,
                full_name=u.full_name or u.email,
                roles=user_roles,
            )
        )
    return out


@router.patch("/users/{user_id}", response_model=UserRead)
def update_user(
    user_id: int,
    payload: UserUpdate,
    request: Request,
    current: User = Depends(require_mutation_permission("user.manage")),
    db: Session = Depends(get_db),
):
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if payload.role_names is not None:
        _ensure_assignable_roles(payload.role_names)
        from app.models.role import Role

        roles = db.scalars(select(Role).where(Role.name.in_(payload.role_names))).all()
        target.roles = list(roles)
    if (
        payload.module_assignments is not None
        or payload.module_access_grants is not None
        or payload.service_access_grants is not None
        or payload.org_capability_grants is not None
        or payload.feature_overrides is not None
        or payload.view_only is not None
    ):
        _apply_access_payload(
            target,
            role_names=target.role_names,
            module_assignments=payload.module_assignments,
            module_access_grants=payload.module_access_grants,
            service_access_grants=payload.service_access_grants,
            org_capability_grants=payload.org_capability_grants,
            feature_overrides=payload.feature_overrides,
            view_only=payload.view_only,
            db=db,
        )
    if payload.region is not None:
        target.region = payload.region
    if payload.is_active is not None:
        target.is_active = payload.is_active
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=current.id, action="update", entity_type="user", entity_id=user_id, **meta)
    db.commit()
    db.refresh(target)
    return _user_to_read(target)


@router.post("/users", response_model=UserRead, status_code=status.HTTP_201_CREATED)
def create_user(
    payload: UserCreate,
    request: Request,
    user: User = Depends(require_mutation_permission("user.manage")),
    db: Session = Depends(get_db),
):
    _ensure_assignable_roles(payload.role_names)
    new_user = auth_service.create_user(
        db,
        email=payload.email,
        password=payload.password,
        full_name=payload.full_name,
        role_names=payload.role_names,
        region=payload.region,
    )
    _apply_access_payload(
        new_user,
        role_names=payload.role_names,
        module_assignments=payload.module_assignments,
        module_access_grants=payload.module_access_grants,
        service_access_grants=payload.service_access_grants,
        org_capability_grants=payload.org_capability_grants,
        feature_overrides=payload.feature_overrides,
        view_only=payload.view_only,
        db=db,
    )
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="create", entity_type="user", entity_id=new_user.id, **meta)
    db.commit()
    db.refresh(new_user)
    return _user_to_read(new_user)


@router.delete("/users/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
def deactivate_user(
    user_id: int,
    request: Request,
    current: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    target.is_active = False
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=current.id, action="deactivate", entity_type="user", entity_id=user_id, **meta)
    db.commit()


@router.post("/therapists/invite")
def invite_therapist(
    payload: InviteCreate,
    request: Request,
    background_tasks: BackgroundTasks,
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    """Legacy generic invite; therapist invites use richer metadata via onboard_therapist_invite."""
    role = payload.role_name or "THERAPIST"
    if role != "THERAPIST":
        _ensure_assignable_roles([role])
    modules = validate_module_assignments([role], payload.module_assignments, db)
    if role == "THERAPIST":
        try:
            display_name = (payload.full_name or "").strip() or payload.email.split("@")[0]
            result = therapist_onboard_svc.onboard_therapist_invite(
                db,
                email=payload.email,
                full_name=display_name,
                phone=None,
                module_assignments=modules,
                services_offered=[],
                short_bio=None,
                created_by_user_id=user.id,
                send_email=payload.send_email,
                background_tasks=background_tasks,
                primary_case_manager_user_id=user.id,
            )
        except ValueError as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
        meta = get_request_meta(request)
        log_audit(
            db,
            actor_user_id=user.id,
            action="invite",
            entity_type="invite_token",
            entity_id=result.get("invite_id"),
            **meta,
        )
        db.commit()
        print(f"[DEV INVITE] {payload.email} -> {result['invite_url']}")
        return {
            "invite_url": result["invite_url"],
            "email": payload.email,
            "expires_at": result["expires_at"],
        }
    token = secrets.token_urlsafe(32)
    invite_meta: dict = {}
    if payload.full_name and payload.full_name.strip():
        invite_meta["full_name"] = payload.full_name.strip()
    if payload.view_only:
        invite_meta["view_only"] = True
    if payload.module_access_grants:
        invite_meta["module_access_grants"] = payload.module_access_grants
    if payload.feature_overrides:
        invite_meta["feature_overrides"] = payload.feature_overrides
    invite = InviteToken(
        email=payload.email.lower(),
        role_name=role,
        module_assignments=modules,
        token=token,
        expires_at=datetime.now(timezone.utc) + timedelta(days=7),
        created_by_user_id=user.id,
        invite_metadata=invite_meta or None,
    )
    db.add(invite)
    db.flush()
    url = f"{settings.frontend_url.rstrip('/')}/invite/{token}"
    if payload.send_email:
        from app.services.email.service import enqueue_portal_invite_email

        display_name = (payload.full_name or "").strip() or payload.email.split("@")[0]
        role_label = role.replace("_", " ").title()
        enqueue_portal_invite_email(
            background_tasks,
            db,
            to=invite.email,
            invite_url=url,
            full_name=display_name,
            role_label=role_label,
            recipient_role=role.lower(),
        )
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="invite", entity_type="invite_token", entity_id=invite.id, **meta)
    db.commit()
    if not payload.send_email:
        print(f"[DEV INVITE] {payload.email} -> {url}")
    return {"invite_url": url, "email": payload.email, "expires_at": invite.expires_at.isoformat()}


@router.post("/therapists/onboard")
def onboard_therapist(
    payload: TherapistOnboardCreate,
    request: Request,
    background_tasks: BackgroundTasks,
    user: User = Depends(require_mutation_permission("user.manage")),
    db: Session = Depends(get_db),
):
    validate_module_assignments(["THERAPIST"], payload.module_assignments, db)
    if payload.services_offered:
        from app.core.therapist_services import validate_service_ids

        try:
            validate_service_ids(payload.services_offered, db)
        except ValueError as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    try:
        result = therapist_onboard_svc.onboard_therapist(
            db,
            email=str(payload.email),
            full_name=payload.full_name,
            phone=payload.phone,
            module_assignments=payload.module_assignments or payload.services_offered,
            services_offered=payload.services_offered,
            mode=payload.mode,
            password=payload.password,
            send_email=payload.send_email,
            short_bio=payload.short_bio,
            created_by_user_id=user.id,
            primary_case_manager_user_id=payload.primary_case_manager_user_id,
            mentor_user_id=payload.mentor_user_id,
            background_tasks=background_tasks,
        )
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="onboard_therapist", entity_type="user", entity_id=result.get("user_id"), **meta)
    db.commit()
    return result


@router.get("/therapists/bulk-template.xlsx")
def therapist_bulk_template(
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    """Return a pre-formatted Excel workbook for bulk therapist upload."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    from app.core.therapist_services import get_service_categories

    svc_cats = get_service_categories(db)
    svc_ids = "|".join(s["id"] for s in svc_cats[:5])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Therapists"

    headers = ["Full Name", "Email", "Phone", "Services (pipe-separated)", "Notes"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = max(18, len(header) + 4)

    ws.append(["Jane Doe", "jane@example.com", "+91 98765 43210", svc_ids, "Example row"])

    svc_ws = wb.create_sheet("Service IDs")
    svc_ws.append(["Service ID", "Label"])
    svc_ws.cell(1, 1).font = Font(bold=True)
    svc_ws.cell(1, 2).font = Font(bold=True)
    for s in svc_cats:
        svc_ws.append([s["id"], s["label"]])
    svc_ws.column_dimensions["A"].width = 28
    svc_ws.column_dimensions["B"].width = 32

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=\"therapist_bulk_template.xlsx\""},
    )


@router.get("/therapists/bulk-template.csv")
def therapist_bulk_template_csv(
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.core.therapist_services import get_service_categories

    svc_cats = get_service_categories(db)
    svc_ids = "|".join(s["id"] for s in svc_cats[:5])
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Full Name", "Email", "Phone", "Services (pipe-separated)", "Notes"])
    writer.writerow(["Jane Doe", "jane@example.com", "+91 98765 43210", svc_ids, "Example row"])
    writer.writerow(["John Smith", "john@example.com", "", "shadow_support", ""])
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=\"therapist_bulk_template.csv\""},
    )


@router.post("/therapists/bulk-onboard", response_model=list[TherapistOnboardResult])
def bulk_onboard_therapists(
    payload: TherapistBulkOnboardRequest,
    request: Request,
    user: User = Depends(require_mutation_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.core.therapist_services import validate_service_ids

    rows = []
    for row in payload.therapists:
        validate_module_assignments(["THERAPIST"], row.module_assignments, db)
        if row.services_offered:
            try:
                validate_service_ids(row.services_offered, db)
            except ValueError as exc:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
        rows.append(row.model_dump())
    results = therapist_onboard_svc.onboard_therapists_bulk(
        db,
        rows,
        mode=payload.mode,
        send_email=payload.send_email,
        created_by_user_id=user.id,
        primary_case_manager_user_id=payload.primary_case_manager_user_id,
        mentor_user_id=payload.mentor_user_id,
    )
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="bulk_onboard_therapists", entity_type="user", entity_id=None, **meta)
    db.commit()
    return [TherapistOnboardResult(**r) for r in results]


@router.get("/therapist-profiles/summary")
def therapist_profiles_summary(
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.models.role import Role

    profiles = profile_svc.list_profiles(db, None)
    therapists = db.scalars(
        select(User).join(User.roles).where(Role.name == RoleName.THERAPIST.value)
    ).all()
    therapist_ids = {t.id for t in therapists}
    profile_user_ids = {p.user_id for p in profiles}
    counts = {"PENDING": 0, "DRAFT": 0, "APPROVED": 0, "PAUSED": 0}
    for p in profiles:
        key = p.status.value if hasattr(p.status, "value") else str(p.status)
        if key in counts:
            counts[key] += 1
    no_profile = len(therapist_ids - profile_user_ids)
    return {**counts, "no_profile": no_profile, "total": len(profiles)}


@router.get("/therapist-profiles", response_model=list[TherapistProfileRead])
def list_therapist_profiles(
    status: Optional[str] = None,
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    st = TherapistProfileStatus(status) if status else None
    profiles = profile_svc.list_profiles(db, st)
    result = []
    for p in profiles:
        u = db.get(User, p.user_id)
        result.append(TherapistProfileRead(**profile_svc.profile_to_dict(p, u)))
    return result


@router.post("/therapist-profiles", response_model=TherapistProfileRead, status_code=status.HTTP_201_CREATED)
def admin_create_therapist_profile(
    payload: TherapistProfileAdminCreate,
    request: Request,
    user: User = Depends(require_mutation_permission("user.manage")),
    db: Session = Depends(get_db),
):
    target = db.get(User, payload.user_id)
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if RoleName.THERAPIST.value not in target.role_names:
        raise HTTPException(status_code=400, detail="User is not a therapist")
    existing = db.scalars(select(TherapistProfile).where(TherapistProfile.user_id == payload.user_id)).first()
    if existing:
        raise HTTPException(status_code=400, detail="Profile already exists for this therapist")
    profile = TherapistProfile(user_id=payload.user_id)
    profile_svc.apply_profile_fields(profile, payload.model_dump(exclude={"user_id", "status"}), db)
    try:
        st = TherapistProfileStatus(payload.status or "APPROVED")
    except ValueError:
        st = TherapistProfileStatus.APPROVED
    profile.status = st
    profile.reviewed_by_user_id = user.id
    profile.reviewed_at = datetime.now(timezone.utc)
    db.add(profile)
    db.flush()
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="create", entity_type="therapist_profile", entity_id=profile.id, **meta)
    db.commit()
    db.refresh(profile)
    return TherapistProfileRead(**profile_svc.profile_to_dict(profile, target))


@router.patch("/therapist-profiles/{profile_id}", response_model=TherapistProfileRead)
def admin_update_therapist_profile(
    profile_id: int,
    payload: TherapistProfileUpdate,
    request: Request,
    user: User = Depends(require_mutation_permission("user.manage")),
    db: Session = Depends(get_db),
):
    profile = db.get(TherapistProfile, profile_id)
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found")
    profile_svc.apply_profile_fields(profile, payload.model_dump(exclude_unset=True), db)
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="update", entity_type="therapist_profile", entity_id=profile_id, **meta)
    db.commit()
    db.refresh(profile)
    target = db.get(User, profile.user_id)
    return TherapistProfileRead(**profile_svc.profile_to_dict(profile, target))


@router.get("/therapist-profiles/{user_id}/reviews", response_model=TherapistReviewsResponse)
def admin_therapist_reviews(
    user_id: int,
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    target = db.get(User, user_id)
    if not target or RoleName.THERAPIST.value not in target.role_names:
        raise HTTPException(status_code=404, detail="Therapist not found")
    rows = review_svc.list_therapist_reviews(db, user_id)
    summary = review_svc.review_summary(db, user_id)
    return TherapistReviewsResponse(
        summary=TherapistReviewSummary(**summary),
        reviews=[TherapistSessionReviewRead(**r) for r in rows],
    )


@router.post("/therapist-profiles/{profile_id}/approve", response_model=TherapistProfileRead)
def admin_approve_profile(
    profile_id: int,
    payload: TherapistProfileReview,
    request: Request,
    user: User = Depends(require_mutation_permission("user.manage")),
    db: Session = Depends(get_db),
):
    profile = db.get(TherapistProfile, profile_id)
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found")
    profile.status = TherapistProfileStatus.APPROVED
    profile.reviewed_by_user_id = user.id
    profile.reviewed_at = datetime.now(timezone.utc)
    if payload.admin_note:
        profile.admin_note = payload.admin_note
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="approve_profile", entity_type="therapist_profile", entity_id=profile_id, **meta)
    db.commit()
    db.refresh(profile)
    return TherapistProfileRead(**profile_svc.profile_to_dict(profile, db.get(User, profile.user_id)))


@router.post("/therapist-profiles/{profile_id}/pause", response_model=TherapistProfileRead)
def admin_pause_profile(
    profile_id: int,
    payload: TherapistProfileReview,
    request: Request,
    user: User = Depends(require_mutation_permission("user.manage")),
    db: Session = Depends(get_db),
):
    profile = db.get(TherapistProfile, profile_id)
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found")
    profile.status = TherapistProfileStatus.PAUSED
    profile.reviewed_by_user_id = user.id
    profile.reviewed_at = datetime.now(timezone.utc)
    if payload.admin_note:
        profile.admin_note = payload.admin_note
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="pause_profile", entity_type="therapist_profile", entity_id=profile_id, **meta)
    db.commit()
    db.refresh(profile)
    return TherapistProfileRead(**profile_svc.profile_to_dict(profile, db.get(User, profile.user_id)))


@router.post("/therapist-profiles/{profile_id}/resume", response_model=TherapistProfileRead)
def admin_resume_profile(
    profile_id: int,
    request: Request,
    user: User = Depends(require_mutation_permission("user.manage")),
    db: Session = Depends(get_db),
):
    profile = db.get(TherapistProfile, profile_id)
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found")
    profile.status = TherapistProfileStatus.APPROVED
    profile.reviewed_by_user_id = user.id
    profile.reviewed_at = datetime.now(timezone.utc)
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="resume_profile", entity_type="therapist_profile", entity_id=profile_id, **meta)
    db.commit()
    db.refresh(profile)
    return TherapistProfileRead(**profile_svc.profile_to_dict(profile, db.get(User, profile.user_id)))


@router.delete("/therapist-profiles/{profile_id}", status_code=status.HTTP_204_NO_CONTENT)
def admin_delete_therapist_profile(
    profile_id: int,
    request: Request,
    user: User = Depends(require_mutation_permission("user.manage")),
    db: Session = Depends(get_db),
):
    profile = db.get(TherapistProfile, profile_id)
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found")
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="delete", entity_type="therapist_profile", entity_id=profile_id, **meta)
    db.delete(profile)
    db.commit()


# --- Case allotment & families ---


@router.get("/cases/next-code")
def admin_next_case_code(
    product_module: str,
    user: User = Depends(require_permission("case.create")),
    db: Session = Depends(get_db),
):
    from app.services import case_code_service

    code = case_code_service.generate_case_code(db, product_module)
    return {"case_code": code, "preview": case_code_service.preview_case_code(product_module)}


@router.get("/cases/pipeline", response_model=AdminCasePipelineBoard)
def admin_cases_pipeline_board(
    user: User = Depends(_admin_dashboard_user),
    db: Session = Depends(get_db),
):
    """Action-oriented case kanban columns (allotment, reassignment, reports, IEP, compliance)."""
    data = case_pipeline_svc.build_pipeline_board(db, user)
    return AdminCasePipelineBoard(**data)


def _iep_reader(user: User = Depends(get_current_user)) -> User:
    if not (
        user_has_permission(user, "attachment.manage")
        or user_has_permission(user, "iep.read")
    ):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient permissions")
    return user


@router.get("/iep/plans")
def admin_list_iep_plans(
    status: Optional[str] = None,
    limit: int = 50,
    user: User = Depends(_iep_reader),
    db: Session = Depends(get_db),
):
    from app.services import iep_plan_service as iep_svc

    return iep_svc.list_plans_scoped(db, user, status=status, limit=min(limit, 100))


@router.get("/iep/dashboard", response_model=AdminIepDashboard)
def admin_iep_dashboard(
    status: Optional[str] = Query(None, description="MISSING | INTERNAL_ONLY | AWAITING_ACK | ACKNOWLEDGED | ALL"),
    product_module: Optional[str] = None,
    search: Optional[str] = None,
    therapist_user_id: Optional[int] = None,
    session_from: Optional[str] = Query(None, description="ISO date YYYY-MM-DD"),
    session_to: Optional[str] = Query(None, description="ISO date YYYY-MM-DD"),
    include_closed: bool = False,
    user: User = Depends(_iep_reader),
    db: Session = Depends(get_db),
):
    data = admin_iep_svc.build_iep_dashboard(
        db,
        user,
        status=status,
        product_module=product_module,
        search=search,
        therapist_user_id=therapist_user_id,
        session_from=session_from,
        session_to=session_to,
        include_closed=include_closed,
    )
    return AdminIepDashboard(**data)


@router.get("/allotment/therapists")
def admin_allotment_therapists(
    product_module: str,
    search: Optional[str] = None,
    approved_only: bool = True,
    user: User = Depends(require_permission("case.assign")),
    db: Session = Depends(get_db),
):
    from app.services import allotment_service

    return allotment_service.list_allotment_therapists(db, user, product_module, search, approved_only)


def _require_family_read(user: User = Depends(get_current_user)) -> User:
    if user_has_permission(user, "user.manage") or user_has_permission(user, "case.create"):
        return user
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Family directory access required")


@router.get("/families")
def admin_list_families(
    search: Optional[str] = None,
    user: User = Depends(_require_family_read),
    db: Session = Depends(get_db),
):
    from app.services import family_admin_service

    return family_admin_service.list_families(db, search)


class ClientBulkRow(BaseModel):
    child_first: str
    child_last: str = ""
    parent_email: str
    parent_full_name: str = ""
    parent_phone: Optional[str] = None


class ClientBulkImportRequest(BaseModel):
    rows: list[ClientBulkRow] = Field(min_length=1, max_length=200)


@router.post("/clients/bulk-import")
def admin_bulk_import_clients(
    payload: ClientBulkImportRequest,
    request: Request,
    user: User = Depends(require_mutation_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.services import family_admin_service

    results = []
    for row in payload.rows:
        try:
            family_admin_service.create_family(
                db,
                parent_email=row.parent_email.strip().lower(),
                parent_full_name=(row.parent_full_name or row.parent_email).strip(),
                parent_phone=row.parent_phone,
                child_first=row.child_first.strip(),
                child_last=(row.child_last or "").strip(),
            )
            results.append({"email": row.parent_email, "success": True, "error": None})
        except Exception as exc:
            results.append({"email": row.parent_email, "success": False, "error": str(exc)})
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="bulk_import_clients", entity_type="child", entity_id=None, **meta)
    db.commit()
    ok = sum(1 for r in results if r["success"])
    return {"total": len(results), "success_count": ok, "results": results}


@router.get("/parents/lookup")
def admin_lookup_parents(
    search: Optional[str] = None,
    user: User = Depends(_require_family_read),
    db: Session = Depends(get_db),
):
    from app.services import family_admin_service

    return family_admin_service.lookup_parents(db, search)


@router.post("/children", status_code=status.HTTP_201_CREATED)
def admin_create_child(
    payload: ChildCreate,
    request: Request,
    user: User = Depends(require_mutation_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.services import family_admin_service

    if not payload.parent_user_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="parent_user_id is required — children must be linked to a parent",
        )
    try:
        result = family_admin_service.add_child_to_parent(
            db,
            parent_user_id=payload.parent_user_id,
            first_name=payload.first_name,
            last_name=payload.last_name,
            date_of_birth=payload.date_of_birth,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="create", entity_type="child", entity_id=result["id"], **meta)
    db.commit()
    return result


@router.post("/families", status_code=status.HTTP_201_CREATED)
def admin_create_family(
    payload: FamilyCreate,
    request: Request,
    user: User = Depends(require_mutation_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.services import family_admin_service

    try:
        result = family_admin_service.create_family(
            db,
            parent_email=payload.parent_email,
            parent_full_name=payload.parent_full_name,
            parent_phone=payload.parent_phone,
            child_first=payload.child.first_name,
            child_last=payload.child.last_name,
            child_dob=payload.child.date_of_birth,
            send_invite=payload.send_invite,
            password=payload.password,
            created_by_user_id=user.id,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="create_family", entity_type="child", entity_id=result["childId"], **meta)
    db.commit()
    if result.get("inviteUrl"):
        print(f"[DEV INVITE] {payload.parent_email} -> {result['inviteUrl']}")
    return result


@router.post("/families/{parent_user_id}/invite")
def admin_invite_parent(
    parent_user_id: int,
    request: Request,
    child_id: Optional[int] = Query(None),
    user: User = Depends(require_mutation_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.services import family_admin_service

    try:
        url = family_admin_service.issue_parent_invite(
            db, parent_user_id, user.id, child_id=child_id
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="invite", entity_type="user", entity_id=parent_user_id, **meta)
    db.commit()
    return {"invite_url": url}


class StatusRequestReview(BaseModel):
    note: Optional[str] = None


@router.get("/status-requests")
def admin_list_status_requests(
    user: User = Depends(require_permission("case.update")),
    db: Session = Depends(get_db),
):
    from app.services import case_status_request_service as csr_svc

    return csr_svc.list_pending(db)


@router.post("/status-requests/{request_id}/approve")
def admin_approve_status_request(
    request_id: int,
    payload: StatusRequestReview,
    request: Request,
    user: User = Depends(require_mutation_permission("case.update")),
    db: Session = Depends(get_db),
):
    from app.models.case_status_request import CaseStatusRequest
    from app.services import case_service, case_status_request_service as csr_svc

    req_row = db.get(CaseStatusRequest, request_id)
    if req_row:
        case = case_service.get_case(db, req_row.case_id)
        if case:
            guard_clinical_case(user, case, db)

    try:
        case = csr_svc.approve_request(db, request_id, user, payload.note)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    meta = get_request_meta(request)
    log_audit(
        db,
        actor_user_id=user.id,
        action="approve_status_request",
        entity_type="case_status_request",
        entity_id=request_id,
        case_id=case.id,
        **meta,
    )
    db.commit()
    return {"status": "approved", "caseId": case.case_code}


@router.post("/status-requests/{request_id}/reject")
def admin_reject_status_request(
    request_id: int,
    payload: StatusRequestReview,
    request: Request,
    user: User = Depends(require_mutation_permission("case.update")),
    db: Session = Depends(get_db),
):
    from app.models.case_status_request import CaseStatusRequest
    from app.services import case_service, case_status_request_service as csr_svc

    if not payload.note or len(payload.note.strip()) < 3:
        raise HTTPException(status_code=400, detail="Review note is required to reject")
    req_row = db.get(CaseStatusRequest, request_id)
    if req_row:
        case = case_service.get_case(db, req_row.case_id)
        if case:
            guard_clinical_case(user, case, db)
    try:
        csr_svc.reject_request(db, request_id, user, payload.note)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    meta = get_request_meta(request)
    log_audit(
        db,
        actor_user_id=user.id,
        action="reject_status_request",
        entity_type="case_status_request",
        entity_id=request_id,
        **meta,
    )
    db.commit()
    return {"status": "rejected"}


@router.get("/observation-checklists")
def admin_list_observation_checklists(
    user: User = Depends(require_permission("monthly_report.approve")),
    db: Session = Depends(get_db),
):
    from app.services import observation_checklist_service as obs_svc

    return obs_svc.list_pending_for_admin(db, user)


@router.get("/observation-checklists/{checklist_id}")
def admin_observation_checklist_detail(
    checklist_id: int,
    user: User = Depends(require_permission("monthly_report.approve")),
    db: Session = Depends(get_db),
):
    from app.models.clinical import ObservationChecklist
    from app.services import case_service, observation_checklist_service as obs_svc

    checklist = db.get(ObservationChecklist, checklist_id)
    if not checklist:
        raise HTTPException(status_code=404, detail="Checklist not found")
    case = case_service.get_case(db, checklist.case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Checklist not found")
    return obs_svc.checklist_to_dict(db, checklist, case, user)


@router.post("/observation-checklists/{checklist_id}/approve")
def admin_approve_observation_checklist(
    checklist_id: int,
    payload: ObservationChecklistReview,
    request: Request,
    user: User = Depends(require_permission("monthly_report.approve")),
    db: Session = Depends(get_db),
):
    from app.models.clinical import ObservationChecklist
    from app.schemas.clinical import ObservationChecklistReview
    from app.services import observation_checklist_service as obs_svc

    from app.services import case_service

    checklist = db.get(ObservationChecklist, checklist_id)
    if not checklist:
        raise HTTPException(status_code=404, detail="Checklist not found")
    case = case_service.get_case(db, checklist.case_id)
    if case:
        guard_clinical_case(user, case, db, feature="reports")
    try:
        obs_svc.approve_checklist(
            db,
            checklist,
            user,
            comment=payload.comment,
            share_with_parent=payload.share_with_parent,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    meta = get_request_meta(request)
    log_audit(
        db,
        actor_user_id=user.id,
        action="approve_observation_checklist",
        entity_type="observation_checklist",
        entity_id=checklist_id,
        case_id=checklist.case_id,
        **meta,
    )
    db.commit()
    if not case:
        case = case_service.get_case(db, checklist.case_id)
    return obs_svc.checklist_to_dict(db, checklist, case, user)


@router.post("/observation-checklists/{checklist_id}/reject")
def admin_reject_observation_checklist(
    checklist_id: int,
    payload: ObservationChecklistReview,
    request: Request,
    user: User = Depends(require_permission("monthly_report.approve")),
    db: Session = Depends(get_db),
):
    from app.models.clinical import ObservationChecklist
    from app.schemas.clinical import ObservationChecklistReview
    from app.services import case_service, observation_checklist_service as obs_svc

    checklist = db.get(ObservationChecklist, checklist_id)
    if not checklist:
        raise HTTPException(status_code=404, detail="Checklist not found")
    case = case_service.get_case(db, checklist.case_id)
    if case:
        guard_clinical_case(user, case, db, feature="reports")
    if not payload.comment or len(payload.comment.strip()) < 3:
        raise HTTPException(status_code=400, detail="Reviewer comment is required")
    try:
        obs_svc.reject_checklist(db, checklist, user, payload.comment)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    meta = get_request_meta(request)
    log_audit(
        db,
        actor_user_id=user.id,
        action="reject_observation_checklist",
        entity_type="observation_checklist",
        entity_id=checklist_id,
        case_id=checklist.case_id,
        **meta,
    )
    db.commit()
    case = case_service.get_case(db, checklist.case_id)
    return obs_svc.checklist_to_dict(db, checklist, case, user)


@router.get("/cases/{case_id}/iep-plans")
def admin_list_iep_plans(
    case_id: int,
    user: User = Depends(require_permission("iep.read")),
    db: Session = Depends(get_db),
):
    from app.services import iep_plan_service as iep_svc

    return iep_svc.list_plans_for_case(db, case_id)


@router.get("/cases/{case_id}/iep-plan")
def admin_get_iep_plan(
    case_id: int,
    user: User = Depends(require_permission("iep.read")),
    db: Session = Depends(get_db),
):
    from app.services import case_service, iep_plan_service as iep_svc

    case = case_service.get_case(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    plan = iep_svc.get_or_create_plan(db, case, user)
    db.commit()
    return iep_svc.plan_to_dict(db, plan, user)


@router.put("/cases/{case_id}/iep-plan")
def admin_save_iep_plan(
    case_id: int,
    payload: IepPlanSave,
    request: Request,
    user: User = Depends(require_mutation_permission("iep.manage")),
    db: Session = Depends(get_db),
):
    from app.services import case_service, iep_plan_service as iep_svc

    case = case_service.get_case(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    guard_clinical_case(user, case, db, feature="iep")
    plan = iep_svc.get_or_create_plan(db, case, user)
    try:
        iep_svc.save_plan(db, plan, payload.sections, user, payload.version)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    meta = get_request_meta(request)
    log_audit(
        db,
        actor_user_id=user.id,
        action="save_iep_plan",
        entity_type="iep_plan",
        entity_id=plan.id,
        case_id=case_id,
        **meta,
    )
    db.commit()
    return iep_svc.plan_to_dict(db, plan, user)


@router.post("/cases/{case_id}/iep-plan/new-version")
def admin_new_iep_plan_version(
    case_id: int,
    request: Request,
    user: User = Depends(require_mutation_permission("iep.manage")),
    db: Session = Depends(get_db),
):
    from app.services import case_service, iep_plan_service as iep_svc

    case = case_service.get_case(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    guard_clinical_case(user, case, db, feature="iep")
    try:
        plan = iep_svc.create_new_version(db, case, user)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    meta = get_request_meta(request)
    log_audit(
        db,
        actor_user_id=user.id,
        action="create_iep_revision",
        entity_type="iep_plan",
        entity_id=plan.id,
        case_id=case_id,
        **meta,
    )
    db.commit()
    return iep_svc.plan_to_dict(db, plan, user)


@router.post("/cases/{case_id}/iep-plan/share-with-parent")
def admin_share_iep_plan(
    case_id: int,
    request: Request,
    user: User = Depends(require_mutation_permission("iep.manage")),
    db: Session = Depends(get_db),
):
    from app.services import case_service, iep_plan_service as iep_svc

    case = case_service.get_case(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    guard_clinical_case(user, case, db, feature="iep")
    plan = iep_svc.get_or_create_plan(db, case, user)
    try:
        iep_svc.share_plan_with_parent(db, plan, user)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    meta = get_request_meta(request)
    log_audit(
        db,
        actor_user_id=user.id,
        action="share_iep_plan",
        entity_type="iep_plan",
        entity_id=plan.id,
        case_id=case_id,
        **meta,
    )
    db.commit()
    return iep_svc.plan_to_dict(db, plan, user)


@router.get("/cases/{case_id}/iep-plan/preview")
def admin_iep_plan_preview(
    case_id: int,
    user: User = Depends(require_permission("iep.read")),
    db: Session = Depends(get_db),
):
    from app.services import case_service, iep_plan_service as iep_svc

    case = case_service.get_case(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    plan = iep_svc.get_latest_plan(db, case_id)
    if not plan:
        raise HTTPException(status_code=404, detail="IEP plan not found")
    html = iep_svc.sections_to_preview_html(db, plan)
    return {"html": html}


@router.get("/cases/{case_id}/iep-plan/export/pdf")
def admin_iep_plan_export_pdf(
    case_id: int,
    user: User = Depends(require_permission("iep.read")),
    db: Session = Depends(get_db),
):
    from fastapi.responses import Response

    from app.services import case_service, iep_plan_service as iep_svc

    case = case_service.get_case(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    plan = iep_svc.get_latest_plan(db, case_id)
    if not plan:
        raise HTTPException(status_code=404, detail="IEP plan not found")
    try:
        data = iep_svc.sections_to_pdf_bytes(db, plan)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF export failed: {e}") from e
    filename = f"IEP_{case.case_code}_{plan.version}.pdf"
    return Response(
        content=data,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/cases/{case_id}/iep-plan/suggestions")
def admin_iep_plan_suggestion(
    case_id: int,
    payload: IepPlanSuggestionCreate,
    user: User = Depends(require_mutation_permission("iep.manage")),
    db: Session = Depends(get_db),
):
    from app.services import case_service, iep_plan_service as iep_svc

    case = case_service.get_case(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    guard_clinical_case(user, case, db, feature="iep")
    plan = iep_svc.get_or_create_plan(db, case, user)
    role = user.role_name if hasattr(user, "role_name") else (user.roles[0].name if user.roles else "ADMIN")
    try:
        row = iep_svc.add_suggestion(db, plan, user, role, payload.body)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    db.commit()
    return iep_svc.plan_to_dict(db, plan, user)


@router.post("/cases/{case_id}/iep-plan/suggestions/resolve")
def admin_resolve_iep_suggestions(
    case_id: int,
    user: User = Depends(require_mutation_permission("iep.manage")),
    db: Session = Depends(get_db),
):
    from app.services import case_service, iep_plan_service as iep_svc

    case = case_service.get_case(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    guard_clinical_case(user, case, db, feature="iep")
    plan = iep_svc.get_latest_plan(db, case_id)
    if not plan:
        raise HTTPException(status_code=404, detail="IEP plan not found")
    iep_svc.resolve_suggestions(db, plan)
    db.commit()
    return iep_svc.plan_to_dict(db, plan, user)


@router.post("/cases/{case_id}/iep-plan/approve")
def admin_approve_iep_plan(
    case_id: int,
    user: User = Depends(require_mutation_permission("iep.manage")),
    db: Session = Depends(get_db),
):
    from app.services import case_service, iep_plan_service as iep_svc

    case = case_service.get_case(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    guard_clinical_case(user, case, db, feature="iep")
    plan = iep_svc.get_latest_plan(db, case_id)
    if not plan:
        raise HTTPException(status_code=404, detail="IEP plan not found")
    iep_svc.approve_plan(db, plan)
    db.commit()
    return iep_svc.plan_to_dict(db, plan, user)


@router.post("/iep/purge-old-versions")
def admin_purge_old_iep_versions(
    case_id: Optional[int] = Query(None),
    user: User = Depends(require_permission("iep.read")),
    db: Session = Depends(get_db),
):
    from app.models.iep_plan import IepPlan
    from app.services import iep_plan_service as iep_svc

    if case_id is not None:
        removed = iep_svc.purge_superseded_iep_plans(db, case_id)
        db.commit()
        return {"caseId": case_id, "removed": removed}
    total = 0
    case_ids = db.scalars(select(IepPlan.case_id).distinct()).all()
    for cid in case_ids:
        total += iep_svc.purge_superseded_iep_plans(db, cid)
    db.commit()
    return {"removed": total}


@router.post("/families/link-by-email")
def admin_link_parent_by_email(
    child_id: int = Query(...),
    parent_email: str = Query(...),
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.services import family_admin_service

    family_admin_service.link_child_to_parent_by_email(db, child_id, parent_email)
    db.commit()
    return {"status": "linked", "child_id": child_id}


@router.post("/families/link")
def admin_link_parent_child(
    parent_user_id: int = Query(...),
    child_id: int = Query(...),
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    from app.models.parent import ParentGuardian
    from app.models.child import Child

    pg = db.scalars(select(ParentGuardian).where(ParentGuardian.user_id == parent_user_id)).first()
    if not pg:
        pg = ParentGuardian(user_id=parent_user_id)
        db.add(pg)
        db.flush()
    child = db.get(Child, child_id)
    if not child:
        raise HTTPException(status_code=404, detail="Child not found")
    if child not in pg.children:
        pg.children.append(child)
    from app.services.parent_service import dedupe_parent_child_links

    dedupe_parent_child_links(db, pg.id)
    db.commit()
    return {"status": "linked"}


@router.post("/cases/allot", status_code=status.HTTP_201_CREATED)
def admin_allot_case(
    payload: CaseAllotRequest,
    request: Request,
    user: User = Depends(require_mutation_permission("case.create")),
    db: Session = Depends(get_db),
):
    from app.services import allotment_service

    data = payload.model_dump()
    ensure_product_module_write_access(user, data.get("product_module") or "homecare", db)
    try:
        result = allotment_service.allot_case(db, user, data)
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    meta = get_request_meta(request)
    log_audit(
        db,
        actor_user_id=user.id,
        action="allot",
        entity_type="case",
        entity_id=result["case"]["id"],
        **meta,
    )
    db.commit()
    return result


def _reports_reader(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> User:
    if not user_has_permission(user, "monthly_report.approve"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient permissions")
    if not user_has_feature(user, "reports", db):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Reports module not enabled")
    return user


def _reports_admin_user(user: User = Depends(_reports_reader)) -> User:
    from app.core.module_access import is_view_only_user

    if is_view_only_user(user):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="View-only access — changes are not allowed")
    return user


def _parse_report_status(value: Optional[str]) -> ReportStatus | None:
    if not value:
        return None
    try:
        return ReportStatus(value.upper())
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid status")


@router.get("/reports/summary", response_model=AdminReportSummary)
def admin_reports_summary(
    user: User = Depends(_reports_reader),
    db: Session = Depends(get_db),
):
    return admin_report_svc.get_summary(db, user)


@router.get("/reports/queue", response_model=PaginatedList[AdminReportListItem])
def admin_reports_queue(
    report_type: Optional[str] = Query(None, alias="type"),
    product_module: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=100),
    user: User = Depends(_reports_reader),
    db: Session = Depends(get_db),
):
    items, meta = admin_report_svc.list_queue_admin(
        db,
        user,
        report_type=report_type,
        product_module=product_module,
        category=category,
        search=search,
        page=page,
        page_size=page_size,
    )
    return PaginatedList[AdminReportListItem](
        items=items,
        total=meta["total"],
        page=meta["page"],
        page_size=meta["page_size"],
        pages=meta["pages"],
    )


@router.get("/reports/monthly", response_model=PaginatedList[AdminReportListItem])
def admin_reports_monthly_list(
    status: Optional[str] = None,
    case_id: Optional[int] = None,
    product_module: Optional[str] = None,
    month: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    parent_review_status: Optional[str] = None,
    queue_only: bool = False,
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=100),
    user: User = Depends(_reports_reader),
    db: Session = Depends(get_db),
):
    items, meta = admin_report_svc.list_monthly_admin(
        db,
        user,
        status=_parse_report_status(status),
        case_id=case_id,
        product_module=product_module,
        month=month,
        category=category,
        search=search,
        parent_review_status=parent_review_status,
        queue_only=queue_only,
        page=page,
        page_size=page_size,
    )
    return PaginatedList[AdminReportListItem](
        items=items,
        total=meta["total"],
        page=meta["page"],
        page_size=meta["page_size"],
        pages=meta["pages"],
    )


@router.get("/reports/observation", response_model=PaginatedList[AdminReportListItem])
def admin_reports_observation_list(
    status: Optional[str] = None,
    case_id: Optional[int] = None,
    product_module: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    queue_only: bool = False,
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=100),
    user: User = Depends(_reports_reader),
    db: Session = Depends(get_db),
):
    items, meta = admin_report_svc.list_observation_admin(
        db,
        user,
        status=_parse_report_status(status),
        case_id=case_id,
        product_module=product_module,
        category=category,
        search=search,
        queue_only=queue_only,
        page=page,
        page_size=page_size,
    )
    return PaginatedList[AdminReportListItem](
        items=items,
        total=meta["total"],
        page=meta["page"],
        page_size=meta["page_size"],
        pages=meta["pages"],
    )


@router.get("/reports/monthly/{report_id}", response_model=AdminReportDetail)
def admin_reports_monthly_detail(
    report_id: int,
    user: User = Depends(_reports_reader),
    db: Session = Depends(get_db),
):
    detail = admin_report_svc.get_monthly_detail(db, user, report_id)
    if not detail:
        raise HTTPException(status_code=404, detail="Report not found")
    return detail


@router.get("/reports/observation/{report_id}", response_model=AdminReportDetail)
def admin_reports_observation_detail(
    report_id: int,
    user: User = Depends(_reports_reader),
    db: Session = Depends(get_db),
):
    detail = admin_report_svc.get_observation_detail(db, user, report_id)
    if not detail:
        raise HTTPException(status_code=404, detail="Report not found")
    return detail


@router.post("/reports/monthly/{report_id}/publish-to-parent", response_model=AdminReportDetail)
def admin_reports_monthly_publish_to_parent(
    report_id: int,
    payload: ReportCommentAction,
    request: Request,
    background_tasks: BackgroundTasks,
    user: User = Depends(_reports_admin_user),
    db: Session = Depends(get_db),
):
    from app.services import case_service, report_service

    report = db.get(MonthlyReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    case = case_service.get_case(db, report.case_id)
    if not case or not case_scope_check(db, user, case):
        raise HTTPException(status_code=404, detail="Report not found")
    guard_clinical_case(user, case, db, feature="reports")
    report_service.sync_published_status(report)
    if report.status != ReportStatus.UNDER_REVIEW:
        raise HTTPException(status_code=400, detail="Report is not awaiting review")
    cm_ready = report_service.user_can_cm_publish(user) and report_service.can_cm_publish(report)
    admin_ready = (
        report_service.user_can_admin_override_publish(user)
        and report_service.can_admin_override_publish(report)
    )
    if cm_ready:
        override = False
    elif admin_ready:
        override = True
    elif report_service.user_can_admin_override_publish(user):
        days = report_service.days_until_admin_override(report)
        raise HTTPException(
            status_code=400,
            detail=f"Admin override available in {days} day(s)" if days is not None else "Admin override not available",
        )
    elif report_service.user_can_cm_publish(user):
        raise HTTPException(status_code=400, detail="Report is not ready for case manager publish")
    else:
        raise HTTPException(status_code=403, detail="Case manager approval required to publish to parents")
    try:
        report_service.publish_monthly_to_parent(
            db,
            report,
            user,
            override=override,
            comment=payload.comment.strip() if payload.comment else None,
        )
        report_service.notify_parents_monthly_report_published(
            db, background_tasks, report=report, case=case
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except PermissionError as e:
        raise HTTPException(status_code=403, detail=str(e))
    meta = get_request_meta(request)
    log_audit(
        db,
        actor_user_id=user.id,
        action="publish_to_parent",
        entity_type="monthly_report",
        entity_id=report_id,
        new_value={"override": override},
        **meta,
    )
    db.commit()
    detail = admin_report_svc.get_monthly_detail(db, user, report_id)
    if not detail:
        raise HTTPException(status_code=404, detail="Report not found")
    return detail


@router.get("/reports/monthly/{report_id}/comments", response_model=list)
def admin_reports_monthly_list_comments(
    report_id: int,
    user: User = Depends(_reports_reader),
    db: Session = Depends(get_db),
):
    from app.services import case_service, report_comment_service

    report = db.get(MonthlyReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    case = case_service.get_case(db, report.case_id)
    if not case or not case_scope_check(db, user, case):
        raise HTTPException(status_code=404, detail="Report not found")
    return report_comment_service.list_monthly_comments(db, report_id)


@router.post("/reports/monthly/{report_id}/comments", response_model=AdminReportDetail)
def admin_reports_monthly_document_comment(
    report_id: int,
    payload: ReportDocumentCommentCreate,
    request: Request,
    user: User = Depends(_reports_admin_user),
    db: Session = Depends(get_db),
):
    from app.services import case_service, report_comment_service

    report = db.get(MonthlyReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    case = case_service.get_case(db, report.case_id)
    if not case or not case_scope_check(db, user, case):
        raise HTTPException(status_code=404, detail="Report not found")
    guard_clinical_case(user, case, db, feature="reports")
    try:
        report_comment_service.add_monthly_comment(
            db, report, user, body=payload.body, comment_type=payload.comment_type
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    meta = get_request_meta(request)
    log_audit(
        db,
        actor_user_id=user.id,
        action="document_comment",
        entity_type="monthly_report",
        entity_id=report_id,
        **meta,
    )
    db.commit()
    detail = admin_report_svc.get_monthly_detail(db, user, report_id)
    if not detail:
        raise HTTPException(status_code=404, detail="Report not found")
    return detail


@router.post("/reports/monthly/{report_id}/cm-review", response_model=AdminReportDetail)
def admin_reports_monthly_cm_review(
    report_id: int,
    payload: CmReviewAction,
    request: Request,
    user: User = Depends(_reports_admin_user),
    db: Session = Depends(get_db),
):
    from app.services import case_service, report_service

    report = db.get(MonthlyReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    case = case_service.get_case(db, report.case_id)
    if not case or not case_scope_check(db, user, case):
        raise HTTPException(status_code=404, detail="Report not found")
    guard_clinical_case(user, case, db, feature="reports")
    if report.status != ReportStatus.UNDER_REVIEW:
        raise HTTPException(status_code=400, detail="Report is not awaiting review")
    if not payload.comment.strip():
        raise HTTPException(status_code=400, detail="Comment is required")
    report_service.cm_review_monthly_report(
        db,
        report,
        user.id,
        comment=payload.comment,
        request_changes=payload.request_changes,
    )
    meta = get_request_meta(request)
    log_audit(
        db,
        actor_user_id=user.id,
        action="cm_review",
        entity_type="monthly_report",
        entity_id=report_id,
        **meta,
    )
    db.commit()
    db.refresh(report)
    detail = admin_report_svc.get_monthly_detail(db, user, report_id)
    if not detail:
        raise HTTPException(status_code=404, detail="Report not found")
    return detail


@router.post("/reports/observation/{report_id}/cm-review", response_model=AdminReportDetail)
def admin_reports_observation_cm_review(
    report_id: int,
    payload: CmReviewAction,
    request: Request,
    user: User = Depends(_reports_admin_user),
    db: Session = Depends(get_db),
):
    from app.services import case_service, report_service

    report = db.get(ObservationReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    case = case_service.get_case(db, report.case_id)
    if not case or not case_scope_check(db, user, case):
        raise HTTPException(status_code=404, detail="Report not found")
    guard_clinical_case(user, case, db, feature="reports")
    if report.status != ReportStatus.UNDER_REVIEW:
        raise HTTPException(status_code=400, detail="Report is not awaiting review")
    if not payload.comment.strip():
        raise HTTPException(status_code=400, detail="Comment is required")
    report_service.cm_review_observation_report(
        db,
        report,
        user.id,
        comment=payload.comment,
        request_changes=payload.request_changes,
    )
    meta = get_request_meta(request)
    log_audit(
        db,
        actor_user_id=user.id,
        action="cm_review",
        entity_type="observation_report",
        entity_id=report_id,
        **meta,
    )
    db.commit()
    detail = admin_report_svc.get_observation_detail(db, user, report_id)
    if not detail:
        raise HTTPException(status_code=404, detail="Report not found")
    return detail


def _report_send_for_review(
    db: Session,
    user: User,
    report_type: str,
    report_id: int,
    payload: SendForReviewAction,
):
    from app.services import case_service, report_service

    if report_type == "monthly":
        report = db.get(MonthlyReport, report_id)
        entity_type = "monthly_report"
    else:
        report = db.get(ObservationReport, report_id)
        entity_type = "observation_report"
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    case = case_service.get_case(db, report.case_id)
    if not case or not case_scope_check(db, user, case):
        raise HTTPException(status_code=404, detail="Report not found")
    guard_clinical_case(user, case, db, feature="reports")
    if report.status not in (ReportStatus.UNDER_REVIEW, ReportStatus.DRAFT, ReportStatus.REJECTED):
        raise HTTPException(status_code=400, detail="Report cannot be sent for review in its current status")
    if not payload.comment.strip():
        raise HTTPException(status_code=400, detail="Comment is required")
    if report_type == "monthly":
        report_service.send_monthly_for_review(
            db,
            report,
            user.id,
            target=payload.target,
            comment=payload.comment,
            case=case,
        )
    else:
        report_service.send_observation_for_review(
            db, report, user.id, target=payload.target, comment=payload.comment
        )
    return entity_type, report


@router.post("/reports/monthly/{report_id}/send-for-review", response_model=AdminReportDetail)
def admin_reports_monthly_send_for_review(
    report_id: int,
    payload: SendForReviewAction,
    request: Request,
    user: User = Depends(_reports_admin_user),
    db: Session = Depends(get_db),
):
    entity_type, _report = _report_send_for_review(db, user, "monthly", report_id, payload)
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="send_for_review", entity_type=entity_type, entity_id=report_id, **meta)
    db.commit()
    detail = admin_report_svc.get_monthly_detail(db, user, report_id)
    if not detail:
        raise HTTPException(status_code=404, detail="Report not found")
    return detail


@router.post("/reports/observation/{report_id}/send-for-review", response_model=AdminReportDetail)
def admin_reports_observation_send_for_review(
    report_id: int,
    payload: SendForReviewAction,
    request: Request,
    user: User = Depends(_reports_admin_user),
    db: Session = Depends(get_db),
):
    entity_type, _report = _report_send_for_review(db, user, "observation", report_id, payload)
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="send_for_review", entity_type=entity_type, entity_id=report_id, **meta)
    db.commit()
    detail = admin_report_svc.get_observation_detail(db, user, report_id)
    if not detail:
        raise HTTPException(status_code=404, detail="Report not found")
    return detail


def _report_add_comment(
    db: Session,
    user: User,
    report_type: str,
    report_id: int,
    payload: ReportCommentAction,
):
    from app.services import case_service, report_service

    if report_type == "monthly":
        report = db.get(MonthlyReport, report_id)
        entity_type = "monthly_report"
    else:
        report = db.get(ObservationReport, report_id)
        entity_type = "observation_report"
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    case = case_service.get_case(db, report.case_id)
    if not case or not case_scope_check(db, user, case):
        raise HTTPException(status_code=404, detail="Report not found")
    guard_clinical_case(user, case, db, feature="reports")
    if not payload.comment.strip():
        raise HTTPException(status_code=400, detail="Comment is required")
    if report_type == "monthly":
        report_service.add_monthly_review_note(db, report, user.id, payload.comment)
    else:
        report_service.add_observation_review_note(db, report, user.id, payload.comment)
    return entity_type


@router.post("/reports/monthly/{report_id}/comment", response_model=AdminReportDetail)
def admin_reports_monthly_comment(
    report_id: int,
    payload: ReportCommentAction,
    request: Request,
    user: User = Depends(_reports_admin_user),
    db: Session = Depends(get_db),
):
    entity_type = _report_add_comment(db, user, "monthly", report_id, payload)
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="review_comment", entity_type=entity_type, entity_id=report_id, **meta)
    db.commit()
    detail = admin_report_svc.get_monthly_detail(db, user, report_id)
    if not detail:
        raise HTTPException(status_code=404, detail="Report not found")
    return detail


@router.post("/reports/observation/{report_id}/comment", response_model=AdminReportDetail)
def admin_reports_observation_comment(
    report_id: int,
    payload: ReportCommentAction,
    request: Request,
    user: User = Depends(_reports_admin_user),
    db: Session = Depends(get_db),
):
    entity_type = _report_add_comment(db, user, "observation", report_id, payload)
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="review_comment", entity_type=entity_type, entity_id=report_id, **meta)
    db.commit()
    detail = admin_report_svc.get_observation_detail(db, user, report_id)
    if not detail:
        raise HTTPException(status_code=404, detail="Report not found")
    return detail


@router.patch("/reports/monthly/{report_id}", response_model=AdminReportDetail)
def admin_reports_monthly_update(
    report_id: int,
    payload: MonthlyReportUpdate,
    request: Request,
    user: User = Depends(_reports_admin_user),
    db: Session = Depends(get_db),
):
    try:
        detail = admin_report_svc.staff_update_monthly(
            db, user, report_id, payload.model_dump(exclude_unset=True)
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if not detail:
        raise HTTPException(status_code=404, detail="Report not found")
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="update", entity_type="monthly_report", entity_id=report_id, **meta)
    db.commit()
    return detail


@router.patch("/reports/observation/{report_id}", response_model=AdminReportDetail)
def admin_reports_observation_update(
    report_id: int,
    payload: ObservationReportUpdate,
    request: Request,
    user: User = Depends(_reports_admin_user),
    db: Session = Depends(get_db),
):
    try:
        detail = admin_report_svc.staff_update_observation(
            db, user, report_id, payload.model_dump(exclude_unset=True)
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if not detail:
        raise HTTPException(status_code=404, detail="Report not found")
    meta = get_request_meta(request)
    log_audit(db, actor_user_id=user.id, action="update", entity_type="observation_report", entity_id=report_id, **meta)
    db.commit()
    return detail


@router.post("/reports/bulk/approve", response_model=BulkReportResult)
def admin_reports_bulk_approve(
    payload: BulkReportAction,
    request: Request,
    user: User = Depends(_reports_admin_user),
    db: Session = Depends(get_db),
):
    result = admin_report_svc.bulk_approve(
        db,
        user,
        report_type=payload.report_type,
        ids=payload.ids,
        comment=payload.comment,
        visibility=payload.visibility_status,
    )
    meta = get_request_meta(request)
    for rid in payload.ids:
        log_audit(
            db,
            actor_user_id=user.id,
            action="bulk_approve",
            entity_type=f"{payload.report_type}_report",
            entity_id=rid,
            **meta,
        )
    db.commit()
    return result


@router.post("/reports/bulk/reject", response_model=BulkReportResult)
def admin_reports_bulk_reject(
    payload: BulkReportAction,
    request: Request,
    user: User = Depends(_reports_admin_user),
    db: Session = Depends(get_db),
):
    result = admin_report_svc.bulk_reject(
        db,
        user,
        report_type=payload.report_type,
        ids=payload.ids,
        comment=payload.comment or "",
    )
    meta = get_request_meta(request)
    for rid in payload.ids:
        log_audit(
            db,
            actor_user_id=user.id,
            action="bulk_reject",
            entity_type=f"{payload.report_type}_report",
            entity_id=rid,
            **meta,
        )
    db.commit()
    return result


@router.get("/reports/export/xlsx")
def admin_reports_export_xlsx(
    report_type: Optional[str] = Query(None, alias="type"),
    status: Optional[str] = None,
    case_id: Optional[int] = None,
    product_module: Optional[str] = None,
    month: Optional[str] = None,
    search: Optional[str] = None,
    queue_only: bool = False,
    user: User = Depends(_reports_reader),
    db: Session = Depends(get_db),
):
    data = admin_report_svc.export_xlsx(
        db,
        user,
        report_type=report_type or "all",
        queue_only=queue_only,
        status=_parse_report_status(status),
        case_id=case_id,
        product_module=product_module,
        month=month,
        search=search,
    )
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="reports-export.xlsx"'},
    )


@router.get("/reports/export/pdf")
def admin_reports_export_pdf(
    report_type: Optional[str] = Query(None, alias="type"),
    status: Optional[str] = None,
    case_id: Optional[int] = None,
    product_module: Optional[str] = None,
    month: Optional[str] = None,
    search: Optional[str] = None,
    queue_only: bool = False,
    user: User = Depends(_reports_reader),
    db: Session = Depends(get_db),
):
    data = admin_report_svc.export_pdf(
        db,
        user,
        report_type=report_type or "all",
        queue_only=queue_only,
        status=_parse_report_status(status),
        case_id=case_id,
        product_module=product_module,
        month=month,
        search=search,
    )
    return Response(
        content=data,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="reports-export.pdf"'},
    )


@router.get("/reports/missing-monthly")
def admin_reports_missing_monthly(
    month: str = Query(..., description="Month label e.g. May 2026"),
    product_module: Optional[str] = None,
    user: User = Depends(_reports_reader),
    db: Session = Depends(get_db),
):
    from app.schemas.report import MissingMonthlyCaseItem

    rows = admin_report_svc.list_missing_monthly(db, user, month=month, product_module=product_module)
    return [MissingMonthlyCaseItem(**r) for r in rows]


@router.get("/invites")
def admin_list_invites(
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    now = datetime.now(timezone.utc)
    rows = db.scalars(
        select(InviteToken)
        .where(InviteToken.used_at.is_(None), InviteToken.expires_at > now)
        .order_by(InviteToken.expires_at.desc())
    ).all()
    return [
        {
            "id": inv.id,
            "email": inv.email,
            "role_name": inv.role_name,
            "module_assignments": inv.module_assignments or [],
            "expires_at": inv.expires_at.isoformat(),
            "invite_url": f"{settings.frontend_url}/invite/{inv.token}",
            "pending_slot_id": (inv.invite_metadata or {}).get("pending_slot_id"),
            "client_name": (inv.invite_metadata or {}).get("client_name"),
            "therapist_user_id": (inv.invite_metadata or {}).get("therapist_user_id"),
        }
        for inv in rows
    ]


@router.post("/invites")
def admin_create_invite(
    payload: InviteCreate,
    request: Request,
    user: User = Depends(require_permission("user.manage")),
    db: Session = Depends(get_db),
):
    return invite_therapist(payload, request, user, db)
